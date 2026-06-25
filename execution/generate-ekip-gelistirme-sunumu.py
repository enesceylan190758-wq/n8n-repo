#!/usr/bin/env python3
"""NefalixAI — Ekip & geliştirme sunumu: ne yaptık, ne yapacağız, kim ne yapar."""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "yonetici-sunum" / "NefalixAI_Ekip_ve_Gelistirme_Plani.xlsx"
TODAY = date(2026, 6, 18)

NAVY, NAVY2, WHITE = "0F2040", "1A2E45", "FFFFFF"
HDR, SEC, GRAY = "0F2040", "E8F4F8", "F5F7FA"
GREEN_L, AMBER_L, RED_L, BLUE_L, TEAL_L = "E8F8EE", "FFF6E6", "FDE8E8", "E8F0FE", "E0F5F2"
DONE, WIP, TODO = "C6EFCE", "FFEB9C", "F2F2F2"
BORDER_C = "D0D7DE"


def border():
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)


def fill(h):
    return PatternFill("solid", fgColor=h)


def fnt(bold=False, size=10, color=NAVY2, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


def col_w(ws, w):
    for c, v in w.items():
        ws.column_dimensions[get_column_letter(c)].width = v


def merge_title(ws, row, title, sub="", cols=8):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=cols)
    c = ws.cell(row, 2, title)
    c.font = fnt(bold=True, size=16, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 30
    if sub:
        r = row + 1
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=cols)
        ws.cell(r, 2, sub).font = fnt(size=10, italic=True)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True)


def section(ws, row, text, cols=8):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=cols)
    c = ws.cell(row, 2, text)
    c.font = fnt(bold=True, size=11, color=NAVY)
    c.fill = fill(SEC)


def hdr(ws, row, headers, start=2):
    for i, h in enumerate(headers, start):
        c = ws.cell(row, i, h)
        c.font = fnt(bold=True, size=10, color=WHITE)
        c.fill = fill(HDR)
        c.border = border()
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 34


def cell(ws, r, c, v, bg=None, bold=False, center=False):
    x = ws.cell(r, c, v)
    x.border = border()
    x.alignment = Alignment(wrap_text=True, vertical="top",
                            horizontal="center" if center else "left")
    if bg:
        x.fill = fill(bg)
    if bold:
        x.font = fnt(bold=True)
    return x


def status_bg(st):
    s = str(st).lower()
    if "tamam" in s or "✓" in s or "canlı" in s:
        return DONE
    if "kısmen" in s or "⏳" in s or "devam" in s or "bekliyor" in s:
        return AMBER_L
    if "yapılacak" in s or "plan" in s or "—" == st:
        return TODO
    if "blok" in s or "askı" in s or "✗" in s:
        return RED_L
    return None


def table(ws, start, headers, rows, widths, status_col=None):
    hdr(ws, start, headers)
    for ri, row in enumerate(rows, start + 1):
        for ci, v in enumerate(row, 2):
            bg = status_bg(v) if status_col and ci == status_col else (GRAY if ri % 2 == 0 else None)
            cell(ws, ri, ci, v, bg=bg)
    col_w(ws, widths)
    return start + len(rows) + 2


def sheet_kapak(wb):
    ws = wb.active
    ws.title = "SUNUM"
    ws.sheet_view.showGridLines = False
    col_w(ws, {1: 2, 2: 18, 3: 18, 4: 18, 5: 18, 6: 2})
    for r in range(1, 35):
        for c in range(1, 7):
            ws.cell(r, c).fill = fill(NAVY)
    ws.merge_cells("B4:E5")
    ws["B4"] = "NEFALIXAI"
    ws["B4"].font = fnt(bold=True, size=28, color=WHITE)
    ws.merge_cells("B6:E7")
    ws["B6"] = "Ekip & Geliştirme Planı\nNe yaptık · Ne yapacağız · Kim ne yapar"
    ws["B6"].font = fnt(size=14, color="CCFBF1")
    ws["B6"].alignment = Alignment(wrap_text=True, horizontal="center")
    boxes = [
        ("B10", "EKİP", "Kurucu + Yazılımcı\n+ AI (Cursor)"),
        ("C10", "PİLOT", "Medident İstanbul"),
        ("D10", "DURUM", "~%55 hazır\nSunucu askıda"),
        ("E10", "HEDEF", "8 klinik · 12 ay"),
    ]
    for ref, t, b in boxes:
        r = int(ref[1:])
        ws.merge_cells(f"{ref[0]}{r}:{ref[0]}{r+2}")
        ws[ref] = f"{t}\n\n{b}"
        ws[ref].font = fnt(size=10, color=WHITE)
        ws[ref].fill = fill("0F766E")
        ws[ref].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells("B14:E16")
    ws["B14"] = (
        f"Toplantı: {TODAY.strftime('%d.%m.%Y')}\n\n"
        "Bu dosya hem yönetim hem yazılımcı için.\n"
        "Teknik detay 'Yazılımcı Kılavuzu' sayfasında."
    )
    ws["B14"].font = fnt(size=11, color="94A3B8")
    ws["B14"].alignment = Alignment(wrap_text=True, horizontal="center")


def sheet_ozet(wb):
    ws = wb.create_sheet("Özet Panel")
    ws.sheet_view.showGridLines = False
    merge_title(ws, 2, "PROJE ÖZET PANELİ", "Tek bakışta ilerleme — sunumun 2. slaytı", 6)
    section(ws, 5, "GENEL İLERLEME", 6)
    hdr(ws, 6, ["Alan", "Durum", "Tamamlanma", "Sıradaki adım"], 2)
    overview = [
        ("Ürün (program özellikleri)", "Kısmen canlı", "%55", "Sunucu aç + Google onay ekranı"),
        ("Klinik paneli (dashboard)", "Canlı (Vercel)", "%70", "Sunucu bağlantısı"),
        ("Veri tabanı (hasta kayıtları)", "Kurulu", "%80", "Sunucuda migration"),
        ("WhatsApp bağlantısı", "Kurulu, QR gerekli", "%40", "QR tara + resmi hat planı"),
        ("Satışa hazırlık", "Plan aşaması", "%25", "Medident case study"),
        ("Yasal (KVKK / İYS)", "Başlanmadı", "%10", "Avukat + metinler"),
        ("Kendi yazılımımıza geçiş", "Plan", "%15", "nefalix-core repo"),
    ]
    for i, row in enumerate(overview, 7):
        for j, v in enumerate(row, 2):
            bg = GREEN_L if "%" in str(v) and int(str(v).replace("%", "")) >= 70 else (
                AMBER_L if "%" in str(v) and int(str(v).replace("%", "")) >= 40 else TODO
            ) if j == 4 else status_bg(row[1] if j == 3 else None)
            cell(ws, i, j, v, bg=bg if j in (3, 4) else GRAY if i % 2 else None)
    col_w(ws, {2: 28, 3: 14, 4: 12, 5: 32})

    section(ws, 16, "EKİP ÖZETİ", 6)
    hdr(ws, 17, ["Rol", "Kim?", "Ne yapar?", "Ne zaman?"], 2)
    team = [
        ("Ürün / karar", "Kurucu (Enes)", "Öncelik, satış, klinik ilişkisi", "Sürekli"),
        ("Yazılım", "Freelance developer (aranacak)", "Panel, API, sunucu", "Temmuz 2026"),
        ("AI destek", "Cursor / AI", "Kod taslağı, script, dokümantasyon", "Sürekli"),
        ("Pilot klinik", "Abdulkadir — Medident", "Günlük kullanım, geri bildirim", "Sunucu sonrası"),
        ("Hukuk", "Avukat (aranacak)", "KVKK, sözleşme, İYS", "Temmuz 2026"),
    ]
    table(ws, 17, ["Rol", "Kim?", "Ne yapar?", "Ne zaman?"], team, {2: 16, 3: 18, 4: 32, 5: 14})


def sheet_yaptik(wb):
    ws = wb.create_sheet("ŞU AN — Ne Yaptık")
    merge_title(ws, 2, "BUGÜNE KADAR TAMAMLANANLAR",
                "Haziran 2026 — somut teslimler", 7)
    rows = [
        ("Ürün vizyonu", "Diş klinikleri için iletişim + itibar programı tanımlandı", "Tamam ✓", "Dokümantasyon"),
        ("Veri modeli", "Hasta mesajı, NPS, yorum, recall tabloları tasarlandı", "Tamam ✓", "Supabase migration"),
        ("8 otomasyon modülü", "Gelen kutusu, NPS, Google, itibar, recall, eNPS, panel, gönderim", "Tamam ✓", "n8n workflow"),
        ("Klinik paneli", "nefalixai.com/dashboard — giriş, KPI, inbox, sekmeler", "Canlı ✓", "Vercel"),
        ("Pilot klinik", "Medident İstanbul profili, logo, linkler", "Tamam ✓", "DB + sync script"),
        ("Sunucu (VPS)", "Türkiye sunucu kuruldu, program yüklendi", "Askıda ⏳", "Kimlik doğrulama"),
        ("WhatsApp altyapısı", "Evolution API kurulumu", "Kısmen ⏳", "QR tarama gerekli"),
        ("Test scriptleri", "Tüm modüller otomatik test", "Tamam ✓", "test-all-workflows.sh"),
        ("Deploy scriptleri", "VPS kurulum, import, aktivasyon", "Tamam ✓", "execution/"),
        ("Yönetim sunumları", "Karar aracı, maliyet, portföy Excel", "Tamam ✓", "docs/yonetici-sunum/"),
        ("Site (landing)", "nefalixai.com tanıtım + chat", "Canlı ✓", "Vercel"),
        ("Pilot kullanıcı", "Abdulkadir panel girişi", "Tamam ✓", "Vercel auth"),
    ]
    table(ws, 5, ["Alan", "Ne yaptık?", "Durum", "Nerede?"], rows,
          {2: 18, 3: 38, 4: 14, 5: 22}, status_col=4)


def sheet_yapacagiz(wb):
    ws = wb.create_sheet("SIRADA — Ne Yapacağız")
    merge_title(ws, 2, "SIRADAKİ İŞLER — ÖNCELİK SIRASIYLA",
                "Yönetici dili + yazılımcı görevi", 8)
    rows = [
        ("P0", "Haz–Tem", "Sunucuyu aç, pilotu çalıştır", "VPS askı kalksın, Medident veri görsün", "Teknik + kurucu", "activate-medident-pilot.sh"),
        ("P0", "Tem", "Medident günlük kullanım", "Abdulkadir haftada 5 gün panele girsin", "Klinik + kurucu", "Eğitim 1 saat"),
        ("P0", "Tem", "KVKK / hukuk paketi", "Satış öncesi yasal metinler", "Avukat", "Sözleşme şablonu"),
        ("P1", "Tem–Ağu", "Google yorum onay ekranı", "Yönetici yorumu onaylayıp yayınlasın", "Yazılımcı", "Dashboard UI"),
        ("P1", "Ağu", "Inbox iyileştirme", "Konuşma görünümü, güvenilir gönderim", "Yazılımcı", "Panel + API"),
        ("P1", "Ağu", "WhatsApp QR + webhook", "Mesajlar panele düşsün", "Teknik", "Evolution"),
        ("P1", "Eyl", "Case study Medident", "Satış için video / PDF", "Kurucu + klinik", "Referans izni"),
        ("P2", "Eki–Ara", "2. ve 3. ücretli klinik", "İlk gelir", "Kurucu satış", "Teşhis aracı Excel"),
        ("P2", "Kas", "nefalix-core API başlangıç", "n8n'den kendi kodumuza geçiş", "Yazılımcı", "Yeni repo"),
        ("P2", "Ara", "5 ücretli klinik", "Satışa hazır kanıt", "Tüm ekip", "MRR hedefi"),
        ("P3", "2027", "Resmi WhatsApp hattı", "Yasal ticari mesaj", "Teknik", "Meta API"),
        ("P3", "2027", "Randevu programı bağlantısı", "HBYS entegrasyonu", "Yazılımcı", "Bulut Klinik vb."),
    ]
    table(ws, 5, ["Öncelik", "Dönem", "İş (sade)", "Bitti sayılır", "Kim?", "Teknik not (yazılımcı)"], rows,
          {2: 8, 3: 10, 4: 24, 5: 28, 6: 16, 7: 24})


def sheet_roller(wb):
    ws = wb.create_sheet("Ekip ve Roller")
    merge_title(ws, 2, "KİM NE YAPAR?", "Çakışma olmasın diye net sınırlar", 7)
    rows = [
        ("Kurucu / CEO", "Enes",
         "Öncelik kararı, satış, klinik ilişkisi, bütçe onayı",
         "Ürün kodu yazmaz; günlük öncelik verir",
         "Haftalık 30 dk ekip toplantısı", "Sürekli"),
        ("Yazılımcı", "Freelance (işe alınacak)",
         "Panel, API, sunucu, test, hata düzeltme",
         "Satış görüşmesi yapmaz; teknik teslim yapar",
         "Haftalık sprint + GitHub", "Tem 2026–"),
        ("AI (Cursor)", "Yapay zeka asistan",
         "Kod taslağı, script, migration, dokümantasyon, hızlı prototip",
         "Tek başına production'a deploy etmez; insan onaylar",
         "Prompt + kod review", "Sürekli"),
        ("Pilot klinik", "Abdulkadir — Medident",
         "Gerçek kullanım, geri bildirim, referans",
         "Teknik kurulum yapmaz",
         "Ayda 1 check-in", "Sunucu sonrası"),
        ("Hukuk", "Avukat",
         "KVKK, sözleşme, İYS, aydınlatma metni",
         "Ürün geliştirmez",
         "Proje bazlı", "Tem 2026"),
    ]
    table(ws, 5, ["Rol", "Kim", "Yapar", "Yapmaz", "Ritim", "Başlangıç"], rows,
          {2: 14, 3: 16, 4: 28, 5: 28, 6: 18, 7: 12})


def sheet_calisma(wb):
    ws = wb.create_sheet("Birlikte Nasıl Çalışırız")
    merge_title(ws, 2, "ÇALIŞMA DÜZENİ",
                "Kurucu + yazılımcı + AI — haftalık ritim", 6)

    section(ws, 5, "HAFTALIK RİTİM", 6)
    rows = [
        ("Pazartesi", "Kurucu", "Haftalık öncelik listesi (max 3 iş)", "Telegram / Notion"),
        ("Salı–Per", "Yazılımcı + AI", "Kod + test + deploy", "GitHub n8n-repo"),
        ("Cuma", "Tüm ekip", "30 dk durum: ne bitti, ne blokaj", "Google Meet"),
        ("Sürekli", "AI (Cursor)", "Script, düzeltme, dokümantasyon", "Cursor IDE"),
        ("Ayda 1", "Medident", "Pilot geri bildirim", "Panel + WhatsApp"),
    ]
    table(ws, 6, ["Ne zaman", "Kim", "Ne", "Araç"], rows, {2: 12, 3: 14, 4: 36, 5: 18})

    section(ws, 13, "KARAR AKIŞI (basit)", 6)
    flow = [
        ("1", "Klinik veya kurucu ihtiyaç söyler", "Örn: Google yorum onayı lazım"),
        ("2", "Kurucu öncelik verir (P0/P1/P2)", "Bu hafta mu? Gelecek ay mı?"),
        ("3", "AI taslak kod / plan yazar", "Cursor + directive"),
        ("4", "Yazılımcı kodlar, test eder, deploy", "PR + test script"),
        ("5", "Pilot klinikte dene", "Medident"),
        ("6", "Çalışıyorsa → satışa aç", "Diğer klinikler"),
    ]
    table(ws, 14, ["Adım", "Ne olur", "Örnek"], flow, {2: 6, 3: 28, 4: 40})

    section(ws, 22, "ARAÇLAR", 6)
    tools = [
        ("Kod", "GitHub — n8n-repo, nefalix-landing", "Tek kaynak"),
        ("Panel", "Vercel — nefalixai.com", "Arayüz"),
        ("Program motoru", "n8n (geçici) → kendi API (hedef)", "İş kuralları"),
        ("Veri", "Supabase — Türkiye VPS", "Hasta verisi"),
        ("AI", "Cursor + OpenAI", "Geliştirme hızı"),
        ("Sunum", "Google Sheets / Excel", "Yönetim"),
    ]
    table(ws, 23, ["Katman", "Araç", "Not"], tools, {2: 14, 3: 32, 4: 28})


def sheet_dev(wb):
    ws = wb.create_sheet("Yazılımcı Kılavuzu")
    merge_title(ws, 2, "YAZILIMCI İÇİN BAŞLANGIÇ KILAVUZU",
                "Bu sayfa teknik ekip içindir — yönetim sunumunda atlanabilir", 7)

    section(ws, 5, "REPOLAR", 7)
    repos = [
        ("n8n-repo", "github.com/.../n8n-repo", "Workflow, DB migration, execution script, VPS"),
        ("nefalix-landing", "Vercel", "Site + dashboard HTML/JS + auth proxy"),
        ("nefalix-core (yakında)", "Yeni repo", "API + servisler — n8n yerine"),
    ]
    table(ws, 6, ["Repo", "Nerede", "İçerik"], repos, {2: 16, 3: 28, 4: 36})

    section(ws, 11, "STACK", 7)
    stack = [
        ("Frontend", "HTML/JS → Next.js (plan)", "nefalix-dashboard.js"),
        ("Backend (şimdi)", "n8n webhooks", "workflows/*.json"),
        ("Backend (hedef)", "Node/TS Fastify", "nefalix-core/"),
        ("DB", "Supabase Postgres", "supabase/migrations/"),
        ("Auth", "Vercel API cookie", "api/auth"),
        ("WhatsApp", "Evolution → Meta API", "docker-compose.evolution.yml"),
        ("AI", "OpenAI gpt-4o-mini", "n8n LangChain nodes"),
        ("Hosting", "Vercel + TR VPS", "docker-compose.prod.yml"),
    ]
    table(ws, 12, ["Katman", "Teknoloji", "Dosya"], stack, {2: 14, 3: 22, 4: 36})

    section(ws, 22, "İLK 2 HAFTA GÖREVLERİ (yazılımcı)", 7)
    tasks = [
        ("1", "Repo clone, .env, local stack", "directives/start_stack.md", "1 gün"),
        ("2", "VPS açılınca activate-medident-pilot.sh", "execution/", "0.5 gün"),
        ("3", "Dashboard → Supabase direct read (n8n bypass)", "wf-07 kapatma hazırlığı", "3 gün"),
        ("4", "Google onay kuyruğu UI", "dashboard + google_reviews tablosu", "4 gün"),
        ("5", "Inbox thread + send fix", "wf-06/08", "3 gün"),
        ("6", "CI: test-all-workflows on push", "GitHub Actions", "1 gün"),
    ]
    table(ws, 23, ["#", "Görev", "Referans", "Süre"], tasks, {2: 4, 3: 32, 4: 28, 5: 10})

    section(ws, 31, "KURALLAR", 7)
    rules = [
        ("Directive oku", "directives/ — iş mantığı burada"),
        ("Script kullan", "execution/ — kendin uydurma"),
        ("Test", "bash execution/test-all-workflows.sh"),
        ("clinic_id sabit", "51738ea8-c12e-40ce-a0e2-42869496d76b"),
        ("Supabase host (Docker)", "host.docker.internal:54321"),
    ]
    for i, (a, b) in enumerate(rules, 32):
        cell(ws, i, 2, a, bold=True)
        cell(ws, i, 3, b)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=6)


def sheet_ai_rol(wb):
    ws = wb.create_sheet("AI Ne Yapar")
    merge_title(ws, 2, "AI (CURSOR) ROLÜ — NET SINIRLAR", "Kurucu ve yazılımcı için", 6)
    rows = [
        ("✓ Yapar", "Workflow ve migration yazma", "Hızlı prototip"),
        ("✓ Yapar", "execution/ script'leri", "Deploy, test, import"),
        ("✓ Yapar", "Directive / dokümantasyon", "SOP güncelleme"),
        ("✓ Yapar", "Hata analizi ve düzeltme önerisi", "Log okuma"),
        ("✓ Yapar", "Excel / sunum üretimi", "Yönetim materyali"),
        ("✗ Yapmaz", "Tek başına production deploy", "İnsan onayı şart"),
        ("✗ Yapmaz", "Klinik satış görüşmesi", "Kurucu işi"),
        ("✗ Yapmaz", "Hukuki karar", "Avukat işi"),
        ("✗ Yapmaz", "WhatsApp QR tarama", "Fiziksel telefon gerekir"),
        ("→ İdeal kullanım", "Yazılımcı prompt yazar → AI kod üretir → yazılımcı review + merge", ""),
    ]
    table(ws, 5, ["Kategori", "Detay", "Not"], rows, {2: 12, 3: 40, 4: 28})


def sheet_mimari(wb):
    ws = wb.create_sheet("Mimari (sade)")
    merge_title(ws, 2, "PROGRAM NASIL ÇALIŞIYOR? — SADE ŞEMA",
                "Yönetici + yeni yazılımcı için", 6)
    section(ws, 5, "BUGÜN", 6)
    cell(ws, 6, 2, (
        "Hasta WhatsApp yazar\n"
        "    ↓\n"
        "Türkiye sunucudaki program (n8n) mesajı alır\n"
        "    ↓\n"
        "Yapay zeka taslak yanıt yazar\n"
        "    ↓\n"
        "Veri tabanına kaydeder\n"
        "    ↓\n"
        "Klinik panelinde (nefalixai.com) sekreter görür ve onaylar"
    ))
    ws.merge_cells("B6:E10")
    ws["B6"].alignment = Alignment(wrap_text=True)
    ws["B6"].fill = fill(BLUE_L)
    ws.row_dimensions[6].height = 100

    section(ws, 12, "HEDEF (12 ay)", 6)
    cell(ws, 13, 2, (
        "Aynı akış — ama ortadaki 'n8n' katmanı\n"
        "kendi yazdığımız API ile değişecek.\n"
        "Panel ve veri tabanı aynı kalır.\n"
        "Klinik fark etmez."
    ))
    ws.merge_cells("B13:E15")
    ws["B13"].fill = fill(GREEN_L)
    ws["B13"].alignment = Alignment(wrap_text=True)


def sheet_sunum_akisi(wb):
    ws = wb.create_sheet("Sunum 25dk")
    merge_title(ws, 2, "EKİP TOPLANTISI — 25 DAKİKA AKIŞI",
                "Kurucu + yazılımcı + varsa klinik — ilk kickoff", 5)
    rows = [
        ("0–3", "SUNUM sayfası", "Proje ne, ekip kim", "Herkes"),
        ("3–8", "ŞU AN — Ne Yaptık", "%55 hazırız, sunucu askı", "Kurucu"),
        ("8–13", "SIRADA — Ne Yapacağız", "P0/P1 listesi", "Kurucu"),
        ("13–18", "Ekip ve Roller + Birlikte Nasıl Çalışırız", "Kim ne yapar", "Herkes"),
        ("18–22", "Yazılımcı Kılavuzu", "Repo, ilk 2 hafta", "Yazılımcı"),
        ("22–25", "Sorular + haftalık Cuma toplantı saati", "", "Herkes"),
    ]
    table(ws, 5, ["Dakika", "Sayfa", "Konu", "Konuşmacı"], rows, {2: 10, 3: 22, 4: 32, 5: 14})


def reorder(wb):
    order = [
        "SUNUM", "Özet Panel", "ŞU AN — Ne Yaptık", "SIRADA — Ne Yapacağız",
        "Ekip ve Roller", "Birlikte Nasıl Çalışırız", "Mimari (sade)",
        "AI Ne Yapar", "Yazılımcı Kılavuzu", "Sunum 25dk",
    ]
    for i, n in enumerate(order):
        wb.move_sheet(n, offset=i - wb.sheetnames.index(n))


def main():
    wb = Workbook()
    sheet_kapak(wb)
    sheet_ozet(wb)
    sheet_yaptik(wb)
    sheet_yapacagiz(wb)
    sheet_roller(wb)
    sheet_calisma(wb)
    sheet_mimari(wb)
    sheet_ai_rol(wb)
    sheet_dev(wb)
    sheet_sunum_akisi(wb)
    reorder(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"✓ Ekip sunumu ({len(wb.sheetnames)} sayfa): {OUT}")


if __name__ == "__main__":
    main()
