#!/usr/bin/env python3
"""NefalixAI — Instagram sosyal medya strateji Excel (10 post).

Usage:
  python3 execution/generate-social-media-excel.py
  python3 execution/generate-social-media-excel.py --output docs/NefalixAI_Sosyal_Medya.xlsx
"""
from __future__ import annotations

import argparse
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "NefalixAI_Sosyal_Medya_Strateji.xlsx"
TODAY = date.today()
MAIL = "nefalixai@gmail.com"
WEB = "nefalixai.com"

C = {
    "navy": "0F172A",
    "gold": "C9A96E",
    "gold_l": "F5F0E6",
    "teal": "0D9488",
    "white": "FFFFFF",
    "slate": "64748B",
    "gray": "F8FAFC",
    "border": "E2E8F0",
    "green_l": "DCFCE7",
}

POSTS = [
    {
        "no": 1,
        "slug": "post_01_brand_hero",
        "type": "Marka / Tanıtım",
        "eyebrow": "YAPAY ZEKÂ DESTEKLİ PLATFORM",
        "headline": "Kliniğinizin Dijital İtibarını\nSiz Uyurken Yönetin.",
        "subtitle": "WhatsApp-first hasta deneyimi, otomatik NPS, Google yorum yönetimi ve itibar koruması — tek panelde, 7/24.",
        "badges": "85+ Klinik Araştırması | KVKK Uyumlu | ~1 Hafta Kurulum",
        "caption": (
            "NefalixAI, klinik ve hizmet işletmeleri için yapay zekâ destekli hasta deneyimi platformudur.\n\n"
            "Tedavi sonrası geri bildirim, Google yorumları, WhatsApp mesajları ve itibar takibi — tek panelde, otomatik.\n\n"
            f"📩 {MAIL}\n🌐 {WEB}"
        ),
        "hashtags": "#NefalixAI #SağlıkTurizmi #KlinikYönetimi #HastaDeneyimi #Dijitalİtibar #WhatsApp #YapayZeka",
        "week": 1,
        "day": "Pazartesi",
        "time": "10:00",
    },
    {
        "no": 2,
        "slug": "post_02_nps_flow",
        "type": "Ürün / NPS",
        "eyebrow": "AKILLI GERİ BİLDİRİM",
        "headline": "Randevu bitti.\n3 saniye sonra hasta zaten cevapladı.",
        "subtitle": "Tedavi sonrası otomatik WhatsApp NPS anketi — memnun hastayı Google'a, memnuniyetsizi size yönlendirir.",
        "badges": "Otomatik Tetikleme | Tek Soru NPS | Akıllı Yönlendirme",
        "caption": (
            "Çoğu klinik tedavi sonrası hastayı takip etmiyor.\n\n"
            "NefalixAI Akıllı Geri Bildirim:\n"
            "✅ Randevu bitince otomatik WhatsApp anketi\n"
            "✅ Tek soru, 3 saniyede yanıt\n"
            "✅ Memnun hasta → Google yorum daveti\n"
            "✅ Memnuniyetsiz hasta → önce siz, sonra çözüm\n\n"
            f"📩 {MAIL}"
        ),
        "hashtags": "#NPS #HastaMemnuniyeti #DişKliniği #GoogleYorumları #NefalixAI",
        "week": 1,
        "day": "Perşembe",
        "time": "18:00",
    },
    {
        "no": 3,
        "slug": "post_03_promoter_rule",
        "type": "İpucu / Değer",
        "eyebrow": "NPS İŞ KURALI",
        "headline": "8–10 puan = Google'a davet\nOlumsuzu internete taşımayın.",
        "subtitle": "Memnun hastayı Google'a yönlendirin; memnuniyetsizliği klinik içinde çözün.",
        "badges": "Promoter → Google | Detractor → Alarm | İtibar Koruması",
        "caption": (
            "Nefalix'te NPS sadece anket değil — akıllı yönlendirme motoru.\n\n"
            "8–10 (promoter): Google yorum linki\n"
            "7 ve altı (detractor): Google istenmez; yöneticiye WhatsApp alarmı\n\n"
            "Sonuç: Daha fazla gerçek yorum, daha az kamuya açık kriz.\n\n"
            f"📩 {MAIL} | {WEB}"
        ),
        "hashtags": "#Googleİşletmem #Onlineİtibar #KlinikPazarlama #NefalixAI",
        "week": 2,
        "day": "Pazartesi",
        "time": "10:00",
    },
    {
        "no": 4,
        "slug": "post_04_sentinel",
        "type": "Premium Modül",
        "eyebrow": "SENTINEL — İTİBAR KORUMA",
        "headline": "Şikayetvar'da yeni şikayet mi?\nYöneticiniz anında haberdar.",
        "subtitle": "Google düşük puan + Şikayetvar tarama → AI duygu analizi → kritik olanlar WhatsApp alarmı.",
        "badges": "Google İzleme | Şikayetvar Sync | Kriz Erken Uyarı",
        "caption": (
            "Olumsuz yorumu geç görmek = geç müdahale.\n\n"
            "Sentinel modülü:\n"
            "🔍 Google düşük puanları izler\n"
            "🔍 Şikayetvar sayfanızı tarar\n"
            "🤖 AI duygu + risk skoru\n"
            "📲 Kritik olanlar yöneticiye WhatsApp özetiyle gider\n\n"
            f"📩 {MAIL}"
        ),
        "hashtags": "#İtibarYönetimi #Şikayetvar #KrizYönetimi #NefalixAI #Sentinel",
        "week": 2,
        "day": "Perşembe",
        "time": "18:00",
    },
    {
        "no": 5,
        "slug": "post_05_review_assistant",
        "type": "Ürün",
        "eyebrow": "YORUM ASİSTANI",
        "headline": "Google'da yeni yorum geldi.\nYanıt taslağı 10 saniyede hazır.",
        "subtitle": "YZ destekli yanıt taslağı — siz onaylarsınız, tek tıkla yayınlanır.",
        "badges": "Duygu Analizi | Marka Tonu | Tek Tık Onay",
        "caption": (
            "Her Google yorumuna saatler harcamak zorunda değilsiniz.\n\n"
            "NefalixAI Yorum Asistanı:\n"
            "• Yeni yorumları otomatik tespit eder\n"
            "• Duygu analizi yapar\n"
            "• Marka tonunuza uygun taslak üretir\n"
            "• Siz onaylarsınız\n\n"
            f"📩 {MAIL}"
        ),
        "hashtags": "#GoogleYorumları #YapayZeka #Klinikİletişimi #NefalixAI",
        "week": 3,
        "day": "Pazartesi",
        "time": "10:00",
    },
    {
        "no": 6,
        "slug": "post_06_inbox",
        "type": "Ürün",
        "eyebrow": "AKILLI MESAJ YÖNETİMİ",
        "headline": "WhatsApp + web sohbet\n= Tek panel.",
        "subtitle": "Tüm kanallar tek gelen kutusunda — AI taslak yanıt, insan onayıyla gönderim.",
        "badges": "Ortak Inbox | AI Taslak | İnsan Onayı",
        "caption": (
            "Resepsiyon WhatsApp'a bakıyor, web sitesi ayrı… Hiçbir mesaj kaybolmamalı.\n\n"
            "NefalixAI Akıllı Mesaj Yönetimi:\n"
            "💬 WhatsApp ve site sohbeti tek gelen kutusunda\n"
            "🤖 AI taslak yanıt hazırlar\n"
            "✅ İnsan onayıyla gönderilir\n\n"
            f"📩 {MAIL} | {WEB}"
        ),
        "hashtags": "#WhatsAppBusiness #KlinikOtomasyon #Hastaİletişimi #NefalixAI",
        "week": 3,
        "day": "Perşembe",
        "time": "18:00",
    },
    {
        "no": 7,
        "slug": "post_07_recall",
        "type": "Premium / ROI",
        "eyebrow": "RECALL — GERİ KAZANIM",
        "headline": "6 aydır gelmeyen hasta\n= Kayıp gelir değil, fırsat.",
        "subtitle": "Uzun süredir gelmeyen hastaları tespit eder, kişiselleştirilmiş WhatsApp hatırlatması gönderir.",
        "badges": "Kayıp Hasta Tespiti | WhatsApp Hatırlatma | Gelir Artışı",
        "caption": (
            "Yeni hasta bulmak, mevcut hastayı geri getirmekten pahalıdır.\n\n"
            "Recall modülü uzun süredir gelmeyen hastaları tespit eder ve kişiselleştirilmiş WhatsApp hatırlatması gönderir.\n\n"
            f"📩 {MAIL}"
        ),
        "hashtags": "#Recall #HastaGeriKazanım #DişKliniği #NefalixAI",
        "week": 4,
        "day": "Pazartesi",
        "time": "10:00",
    },
    {
        "no": 8,
        "slug": "post_08_enps",
        "type": "Farklılaşma",
        "eyebrow": "ÇALIŞAN DENEYİMİ",
        "headline": "Hasta memnuniyeti kadar\nekip nabzı da önemli.",
        "subtitle": "Haftalık eNPS nabız anketi, trend takibi ve aksiyon maddeleri dashboard'da.",
        "badges": "Haftalık eNPS | Trend Takibi | Aksiyon Listesi",
        "caption": (
            "İtibar sadece hastadan gelmez — ekibinizden de gelir.\n\n"
            "NefalixAI Çalışan Deneyimi:\n"
            "📊 Haftalık eNPS nabız anketi\n"
            "📈 Trend takibi\n"
            "✅ Aksiyon maddeleri dashboard'da\n\n"
            f"📩 {MAIL}"
        ),
        "hashtags": "#eNPS #KlinikYönetimi #NefalixAI",
        "week": 4,
        "day": "Perşembe",
        "time": "18:00",
    },
    {
        "no": 9,
        "slug": "post_09_kvkk",
        "type": "Güven",
        "eyebrow": "GÜVEN & KVKK",
        "headline": "Hasta verisi ciddi iş.\nKVKK uyumlu altyapı.",
        "subtitle": "YZ yanıtları insan onayından geçer; veri güvenliği önceliklidir.",
        "badges": "KVKK Uyumlu | İnsan Onayı | Veri Güvenliği",
        "caption": (
            "Sağlık verisiyle çalışan bir platformda hız, güvenin önüne geçmemeli.\n\n"
            "NefalixAI:\n"
            "🔒 KVKK uyumlu veri mimarisi\n"
            "🔒 YZ yanıtları insan onayından geçer\n"
            "🔒 Ticari WhatsApp öncesi İYS izin kontrolü\n\n"
            f"📩 {MAIL} | {WEB}/kvkk"
        ),
        "hashtags": "#KVKK #SağlıkTeknolojisi #VeriGüvenliği #NefalixAI",
        "week": 5,
        "day": "Pazartesi",
        "time": "10:00",
    },
    {
        "no": 10,
        "slug": "post_10_demo_cta",
        "type": "CTA / Dönüşüm",
        "eyebrow": "CANLI DEMO",
        "headline": "15 dakikada canlı demo.\nKliniğinize özel akışı kuralım.",
        "subtitle": "6 modül, tek omurga — kurulum ~1 hafta, pilot kliniklerle sahadan doğrulanmış.",
        "badges": "6 Modül | ~1 Hafta Kurulum | Canlı Demo",
        "caption": (
            "NefalixAI 6 modül, tek omurga:\n"
            "⭐ Akıllı Geri Bildirim\n"
            "💬 Mesaj Yönetimi\n"
            "🛡️ Yorum Asistanı\n"
            "👥 Çalışan Deneyimi\n"
            "🔍 Sentinel\n"
            "🔁 Recall\n\n"
            f"Demo için yazın:\n📩 {MAIL}\n🌐 {WEB}"
        ),
        "hashtags": "#NefalixAI #Demo #SağlıkTurizmi #KlinikDijitalleşme #WhatsApp",
        "week": 5,
        "day": "Perşembe",
        "time": "18:00",
    },
]


def _border():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(h: str):
    return PatternFill("solid", fgColor=h)


def _font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


def col_w(ws, widths: dict[int, float]):
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def sheet_cover(wb: Workbook):
    ws = wb.active
    ws.title = "Kapak"
    ws.sheet_view.showGridLines = False
    col_w(ws, {1: 4, 2: 50, 3: 4})

    for r in range(1, 28):
        ws.row_dimensions[r].height = 22
        for c in (1, 2, 3):
            ws.cell(r, c).fill = _fill(C["navy"])

    ws.merge_cells("B3:B3")
    ws["B3"] = "NEFALIXAI"
    ws["B3"].font = _font(bold=True, size=32, color=C["gold"])
    ws["B4"] = "Instagram Sosyal Medya Stratejisi"
    ws["B4"].font = _font(size=16, color=C["white"])
    ws["B6"] = "10 Post | Marka + Ürün + Güven | Manuel planlayıcı yükleme"
    ws["B6"].font = _font(size=12, color=C["gold_l"])

    info = [
        ("Platform", "Instagram (manuel planlayıcı)"),
        ("Görsel boyut", "1080 × 1080 px"),
        ("Üretim", "GPT görsel + Drive paket (otomasyon)"),
        ("İletişim", MAIL),
        ("Web", WEB),
        ("Tel", "YOK — yazılmaz"),
        ("Tarih", TODAY.strftime("%d.%m.%Y")),
    ]
    r = 9
    for label, val in info:
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = _font(bold=True, size=11, color=C["gold"])
        ws.merge_cells(f"C{r}:C{r}")  # noop guard
        ws.cell(r, 2, f"{label}:  {val}").font = _font(size=11, color=C["white"])
        r += 1

    ws["B18"] = "Sayfalar: Kapak → 10 Post → Takvim → Marka Kuralları → Görsel Brief"
    ws["B18"].font = _font(size=10, color=C["slate"], italic=True)


def sheet_posts(wb: Workbook):
    ws = wb.create_sheet("10 Post")
    headers = [
        "#", "Tür", "Hafta", "Gün", "Saat", "Eyebrow", "Görsel Başlık", "Alt Başlık",
        "Badge'ler", "Instagram Caption", "Hashtag", "Slug", "Durum",
    ]
    col_w(ws, {1: 4, 2: 14, 3: 6, 4: 10, 5: 8, 6: 22, 7: 32, 8: 36, 9: 28, 10: 48, 11: 36, 12: 22, 13: 12})

    ws.merge_cells("A1:M1")
    ws["A1"] = "10 POST İÇERİK PLANI — NefalixAI"
    ws["A1"].font = _font(bold=True, size=16, color=C["navy"])

    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.font = _font(bold=True, size=10, color=C["white"])
        c.fill = _fill(C["navy"])
        c.border = _border()
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for ri, p in enumerate(POSTS, 4):
        row = [
            p["no"], p["type"], p["week"], p["day"], p["time"], p["eyebrow"],
            p["headline"], p["subtitle"], p["badges"], p["caption"], p["hashtags"],
            p["slug"], "Taslak",
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.border = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if ri % 2 == 0:
                cell.fill = _fill(C["gray"])
        ws.row_dimensions[ri].height = 120


def sheet_calendar(wb: Workbook):
    ws = wb.create_sheet("Takvim")
    col_w(ws, {1: 6, 2: 8, 3: 14, 4: 12, 5: 10, 6: 28, 7: 14, 8: 20})
    ws["A1"] = "YAYIN TAKVİMİ (5 hafta × 2 post)"
    ws["A1"].font = _font(bold=True, size=14, color=C["navy"])

    headers = ["Hafta", "Post #", "Gün", "Saat", "Konu", "Platform", "Not"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.font = _font(bold=True, color=C["white"])
        c.fill = _fill(C["teal"])
        c.border = _border()

    for ri, p in enumerate(POSTS, 4):
        ws.cell(ri, 1, f"Hafta {p['week']}").border = _border()
        ws.cell(ri, 2, p["no"]).border = _border()
        ws.cell(ri, 3, p["day"]).border = _border()
        ws.cell(ri, 4, p["time"]).border = _border()
        ws.cell(ri, 5, p["eyebrow"]).border = _border()
        ws.cell(ri, 6, "Instagram").border = _border()
        ws.cell(ri, 7, "Drive'dan görsel + caption yapıştır").border = _border()
        for c in range(1, 8):
            ws.cell(ri, c).alignment = Alignment(wrap_text=True, vertical="top")


def sheet_brand(wb: Workbook):
    ws = wb.create_sheet("Marka Kuralları")
    col_w(ws, {1: 22, 2: 70})
    rules = [
        ("Marka adı", "NEFALIXAI — NEFALIX lacivert/beyaz, AI altın"),
        ("Renkler", "Arka plan #0D1B2A / #0A0F1C | Vurgu altın #C9A96E | Metin beyaz"),
        ("Boyut", "1080×1080 px kare"),
        ("Mail", MAIL + " — her postta"),
        ("Web", WEB),
        ("Telefon", "ASLA yazma"),
        ("Ton", "Premium sağlık-tech, kurumsal, güven veren, Türkçe"),
        ("Logo", "Şeffaf PNG, sağ üst veya sol üst, minimum boşluk"),
        ("Yasak", "Stok insan yüzü, yanlış marka, telefon, neon renk"),
        ("Üretim", "python3 execution/social-generate-next.py → Drive klasörü"),
        ("Yükleme", "Instagram planlayıcı — manuel"),
    ]
    ws["A1"] = "MARKA & İLETİŞİM KURALLARI"
    ws["A1"].font = _font(bold=True, size=14, color=C["navy"])
    for i, (k, v) in enumerate(rules, 3):
        ws.cell(i, 1, k).font = _font(bold=True)
        ws.cell(i, 1).border = _border()
        ws.cell(i, 1).fill = _fill(C["gold_l"])
        ws.cell(i, 2, v).border = _border()
        ws.cell(i, 2).alignment = Alignment(wrap_text=True)


def sheet_visual_brief(wb: Workbook):
    ws = wb.create_sheet("Görsel Brief")
    col_w(ws, {1: 4, 2: 20, 3: 55, 4: 40})
    ws["A1"] = "MANUS / GPT GÖRSEL PROMPT ŞABLONU"
    ws["A1"].font = _font(bold=True, size=14, color=C["navy"])

    template = """Design a premium Instagram square post (1080x1080).

BRAND: NefalixAI — health-tech, navy #0D1B2A + gold #C9A96E
Eyebrow: {eyebrow}
Headline: {headline}
Subtitle: {subtitle}
Badges: {badges}
Footer: {mail} | {web}
Post {no}/10

RULES: No phone numbers. Turkish text readable. Luxury medical SaaS style."""

    ws.merge_cells("B3:D3")
    ws["B3"] = "Genel şablon (her post için doldur)"
    ws["B3"].font = _font(bold=True)
    ws.merge_cells("B4:D8")
    ws["B4"] = template.format(
        eyebrow="{eyebrow}", headline="{headline}", subtitle="{subtitle}",
        badges="{badges}", mail=MAIL, web=WEB, no="{no}",
    )
    ws["B4"].alignment = Alignment(wrap_text=True)
    ws["B4"].fill = _fill(C["gray"])

    r = 10
    ws.cell(r, 2, "Post").font = _font(bold=True)
    ws.cell(r, 3, "Hazır Prompt").font = _font(bold=True)
    ws.cell(r, 4, "Caption dosyası").font = _font(bold=True)
    for c in range(2, 5):
        ws.cell(r, c).fill = _fill(C["navy"])
        ws.cell(r, c).font = _font(bold=True, color=C["white"])
        ws.cell(r, c).border = _border()

    for p in POSTS:
        r += 1
        prompt = template.format(
            eyebrow=p["eyebrow"],
            headline=p["headline"].replace("\n", " / "),
            subtitle=p["subtitle"],
            badges=p["badges"],
            mail=MAIL,
            web=WEB,
            no=p["no"],
        )
        ws.cell(r, 2, f"#{p['no']}").border = _border()
        ws.cell(r, 3, prompt).border = _border()
        ws.cell(r, 3).alignment = Alignment(wrap_text=True)
        ws.cell(r, 4, f"{p['slug']}_caption.txt").border = _border()
        ws.row_dimensions[r].height = 80


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=OUT)
    args = parser.parse_args()

    wb = Workbook()
    sheet_cover(wb)
    sheet_posts(wb)
    sheet_calendar(wb)
    sheet_brand(wb)
    sheet_visual_brief(wb)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"✓ Excel: {args.output} ({len(wb.sheetnames)} sayfa, {len(POSTS)} post)")


if __name__ == "__main__":
    main()
