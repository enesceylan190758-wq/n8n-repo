#!/usr/bin/env python3
"""Nefalix SaaS geliştirme matrisi — nasıl yapacağız?"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "Nefalix_SaaS_Gelistirme_Matrisi.xlsx"
TODAY = date(2026, 6, 18)

C = {
    "navy": "0F172A", "teal": "0D9488", "teal_l": "CCFBF1", "teal_d": "0F766E",
    "white": "FFFFFF", "slate": "64748B", "gray": "F8FAFC", "border": "E2E8F0",
    "green_l": "DCFCE7", "green": "16A34A", "amber_l": "FEF3C7", "amber": "F59E0B",
    "red_l": "FEE2E2", "blue_l": "DBEAFE", "purple_l": "EDE9FE",
    "n8n": "FEF3C7", "api": "DBEAFE", "done": "DCFCE7", "todo": "F1F5F9",
}


def _border():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(h):
    return PatternFill("solid", fgColor=h)


def _f(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)


def col_w(ws, w: dict):
    for c, v in w.items():
        ws.column_dimensions[get_column_letter(c)].width = v


def title(ws, r, text, sub=""):
    ws.merge_cells(f"A{r}:H{r}")
    ws[f"A{r}"] = text
    ws[f"A{r}"].font = _f(bold=True, size=18, color=C["navy"])
    if sub:
        ws.merge_cells(f"A{r+1}:H{r+1}")
        ws[f"A{r+1}"] = sub
        ws[f"A{r+1}"].font = _f(size=10, color=C["slate"])


def hdr(ws, row, headers, bg=C["teal"]):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = _f(bold=True, size=10, color=C["white"])
        c.fill = _fill(bg)
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def cell(ws, r, c, val, bg=None, bold=False, center=False):
    x = ws.cell(r, c, val)
    x.border = _border()
    x.alignment = Alignment(wrap_text=True, vertical="top",
                            horizontal="center" if center else "left")
    if bg:
        x.fill = _fill(bg)
    if bold:
        x.font = _f(bold=True)
    return x


def status_color(st: str) -> str:
    s = str(st).lower()
    if "canlı" in s or "tamam" in s or "✓" in s:
        return C["done"]
    if "n8n" in s or "pilot" in s or "geçiş" in s:
        return C["n8n"]
    if "api" in s or "geliştir" in s or "dev" in s:
        return C["api"]
    if "plan" in s or "⏳" in s:
        return C["todo"]
    return C["gray"]


# ═══════════════════════════════════════════════════════════════
# 0 — MASTER: NASIL YAPACAĞIZ?
# ═══════════════════════════════════════════════════════════════
def sheet_master(wb):
    ws = wb.active
    ws.title = "0-Nasıl Yapacağız"
    ws.sheet_view.showGridLines = False
    col_w(ws, {1: 3, 2: 22, 3: 22, 4: 22, 5: 22, 6: 3})

    title(ws, 1, "NEFALIX SAAS — GELİŞTİRME MATRİSİ",
          f"Otomasyondan yazılıma geçiş | {TODAY.strftime('%d.%m.%Y')}")

    # 4 sütunlu strateji
    blocks = [
        ("B4", C["teal_d"], "1. VERİ ÖNCE",
         "Supabase = tek doğruluk kaynağı.\nTüm modüller buraya yazar/okur.\nDashboard doğrudan DB'ye bağlanacak."),
        ("C4", C["navy"], "2. MODÜL MODÜL",
         "Tek dev kod yok.\nHer özellik ayrı servis.\nBiri bozulunca diğeri çalışır."),
        ("D4", C["teal_d"], "3. STRANGLER",
         "n8n önce çalışır.\nHer modül hazır olunca\nkendi API'mize taşınır."),
        ("E4", C["navy"], "4. ÖNCE DEĞER",
         "Inbox + Google onay →\nsonra grafik →\nsonra HBYS."),
    ]
    for ref, bg, t, body in blocks:
        c0, r0 = ref[0], int(ref[1:])
        ws.merge_cells(f"{c0}{r0}:{c0}{r0+4}")
        ws[ref] = f"{t}\n\n{body}"
        ws[ref].font = _f(size=10, color=C["white"])
        ws[ref].fill = _fill(bg)
        ws[ref].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r0].height = 90

    r = 11
    ws.merge_cells(f"B{r}:E{r}")
    ws[f"B{r}"] = "5 KATMAN — HER KATMANIN TEK SORUMLUSU"
    ws[f"B{r}"].font = _f(bold=True, size=13, color=C["navy"])
    r += 1
    hdr(ws, r, ["Katman", "Ne?", "Bugün", "Hedef", "Araç"])
    layers = [
        ("① Ekran", "Klinik paneli", "HTML/JS + Vercel", "Next.js + React", "Vercel"),
        ("② API", "İş kuralları kapısı", "n8n webhook", "Node/TS REST API", "VPS veya Vercel"),
        ("③ İş mantığı", "Kurallar + AI prompt", "n8n workflow", "TypeScript servisler", "Repo: nefalix-core"),
        ("④ Veri", "Hasta, inbox, NPS…", "Supabase TR", "Aynı + RLS", "Postgres VPS"),
        ("⑤ Entegrasyon", "WA, HBYS, Google", "Evolution + webhook", "Meta API + partner", "Resmi API'ler"),
    ]
    for i, row in enumerate(layers, r + 1):
        for j, v in enumerate(row, 1):
            bg = C["teal_l"] if j == 1 else (C["gray"] if i % 2 == 0 else None)
            cell(ws, i, j, v, bg=bg, bold=(j == 1))

    r = 20
    ws.merge_cells(f"B{r}:E{r}")
    ws[f"B{r}"] = "GELİŞTİRME DÖNGÜSÜ (her modül için tekrarla)"
    ws[f"B{r}"].font = _f(bold=True, size=13, color=C["navy"])
    r += 1
    cycle = [
        "① Tanımla → n8n workflow'unu dokümante et (girdi/çıktı tablo)",
        "② Kodla → TypeScript servis + test yaz",
        "③ API → /api/v1/... endpoint aç",
        "④ UI → Dashboard'da ekran bağla",
        "⑤ Test → smoke test + Medident'te pilot",
        "⑥ Kapat → n8n workflow'unu deaktif et",
    ]
    for i, step in enumerate(cycle, r):
        ws.merge_cells(f"B{i}:E{i}")
        cell(ws, i, 2, step, bg=C["gray"] if i % 2 else C["blue_l"])


# ═══════════════════════════════════════════════════════════════
# 1 — MODÜL × YOL MATRİSİ (ana matris)
# ═══════════════════════════════════════════════════════════════
def sheet_module_matrix(wb):
    ws = wb.create_sheet("1-Modül Matrisi")
    title(ws, 1, "MODÜL × GELİŞTİRME YOLU", "Her satır = bir ürün özelliği | Sütunlar = nasıl yapılacak")
    col_w(ws, {1: 16, 2: 10, 3: 12, 4: 14, 5: 14, 6: 14, 7: 12, 8: 12, 9: 14, 10: 10, 11: 12})

    headers = [
        "Modül", "Öncelik", "Bugün\n(durum)", "Veri tablosu",
        "Tetikleyici", "n8n\n(şimdi)", "API\n(hedef)", "UI\n(hedef)",
        "Entegrasyon", "Faz", "Tahmini\nsüre",
    ]
    hdr(ws, 3, headers)

    modules = [
        ("Dashboard okuma", "P0", "n8n webhook", "clinics, nps, inbox…", "GET istek",
         "wf-07", "GET /api/v1/dashboard", "Mevcut panel", "Supabase", "F0", "1 hf"),
        ("Inbox gelen", "P0", "n8n wf-06", "inbox_messages", "WA webhook",
         "wf-06", "POST /inbox/incoming", "Inbox sekmesi", "Evolution→API", "F1", "2 hf"),
        ("Inbox gönder", "P0", "n8n wf-08", "inbox_messages", "Panel buton",
         "wf-08", "POST /inbox/send", "Modal gönder", "Evolution/Meta", "F1", "1 hf"),
        ("Google yorum AI", "P1", "n8n wf-02", "google_reviews", "Webhook/cron",
         "wf-02", "POST /reviews/analyze", "Onay kuyruğu", "OpenAI", "F1", "2 hf"),
        ("NPS / Feedback", "P1", "n8n wf-01", "nps_responses", "HBYS webhook",
         "wf-01", "POST /nps/response", "NPS sekmesi", "WhatsApp", "F2", "2 hf"),
        ("İtibar Sentinel", "P2", "n8n wf-04", "reputation_mentions", "Webhook",
         "wf-04", "POST /sentinel/mention", "Sentinel sekmesi", "OpenAI", "F2", "1 hf"),
        ("Recall", "P2", "n8n wf-05", "recall_campaigns", "Cron",
         "wf-05", "POST /recall/run", "Recall sekmesi", "WhatsApp+İYS", "F3", "2 hf"),
        ("eNPS", "P3", "n8n wf-03", "enps_responses", "Webhook",
         "wf-03", "POST /enps/response", "eNPS sekmesi", "—", "F3", "1 hf"),
        ("Web chatbot", "P1", "n8n chatbot", "inbox_messages", "Site widget",
         "chat wf", "POST /chat/message", "Site embed", "OpenAI", "F2", "2 hf"),
        ("Auth / roller", "P0", "Vercel API", "—", "Login",
         "api/auth", "JWT + RLS", "Login sayfası", "Supabase Auth", "F0", "1 hf"),
        ("Klinik ayarları", "P2", "Yok", "clinics", "Panel",
         "—", "PATCH /clinics/:id", "Ayarlar sayfası", "—", "F3", "2 hf"),
        ("Bildirimler", "P2", "Yok", "—", "Event",
         "—", "Queue worker", "E-posta/WA", "Resend", "F3", "2 hf"),
        ("HBYS sync", "P3", "Webhook stub", "patients, appointments", "HBYS",
         "router", "POST /hbys/*", "Entegrasyon UI", "Bulut Klinik API", "F4", "4 hf"),
    ]

    for ri, row in enumerate(modules, 4):
        for ci, val in enumerate(row, 1):
            bg = None
            if ci == 3:
                bg = status_color(str(val))
            elif ci == 6:
                bg = C["n8n"]
            elif ci == 7:
                bg = C["api"]
            elif ci == 2:
                bg = C["red_l"] if val == "P0" else (C["amber_l"] if val == "P1" else C["gray"])
            cell(ws, ri, ci, val, bg=bg, center=(ci in (2, 10, 11)))

    r = 4 + len(modules) + 2
    ws.merge_cells(f"A{r}:K{r}")
    ws[f"A{r}"] = "Renk: sarı = n8n'de çalışıyor | mavi = API'ye taşınacak | yeşil = canlı"
    ws[f"A{r}"].font = _f(size=9, color=C["slate"])


# ═══════════════════════════════════════════════════════════════
# 2 — FAZ × ÇIKTI MATRİSİ
# ═══════════════════════════════════════════════════════════════
def sheet_phase_matrix(wb):
    ws = wb.create_sheet("2-Faz Matrisi")
    title(ws, 1, "FAZ × ÇIKTI × SAHİP", "Ne zaman ne teslim edilir")
    col_w(ws, {1: 8, 2: 14, 3: 28, 4: 28, 5: 12, 6: 12, 7: 14})

    hdr(ws, 3, ["Faz", "Dönem", "Teslim (ürün)", "Teslim (teknik)", "Sahip", "Süre", "Bitti kriteri"])
    phases = [
        ("F0", "Haz-Tem 26", "Panel veri gelir, sağlık göstergesi",
         "VPS + n8n aktif + credential", "Teknik", "4 hf",
         "Medident 5 gün/hafta panel"),
        ("F1", "Ağu-Eyl 26", "Inbox thread + Google onay + WA gönder",
         "API v1: inbox + reviews; n8n wf-06/08 kapat", "Dev", "6 hf",
         "%80 yanıt panelden"),
        ("F2", "Eki-Kas 26", "Grafikler, mobil, PDF rapor",
         "Next.js migrasyon başlangıç; dashboard→Supabase direct", "Dev", "6 hf",
         "Mobil inbox çalışır"),
        ("F3", "Ara 26-Mar 27", "5 ödeyen klinik, çoklu klinik, ayarlar",
         "Auth RLS, İYS middleware, billing stub", "Ürün+Satış", "12 hf",
         "5 fatura kesiliyor"),
        ("F4", "Nis-Ağu 27", "HBYS stub, resmi WhatsApp",
         "n8n tamamen kapat; queue worker", "Dev", "16 hf",
         "n8n container durduruldu"),
    ]
    for ri, row in enumerate(phases, 4):
        for ci, v in enumerate(row, 1):
            cell(ws, ri, ci, v, bg=C["teal_l"] if ci == 1 else (C["gray"] if ri % 2 == 0 else None), bold=(ci == 1))


# ═══════════════════════════════════════════════════════════════
# 3 — TEKNOLOJİ SEÇİM MATRİSİ
# ═══════════════════════════════════════════════════════════════
def sheet_tech_matrix(wb):
    ws = wb.create_sheet("3-Teknoloji")
    title(ws, 1, "TEKNOLOJİ SEÇİM MATRİSİ", "Neyi neden seçiyoruz — alternatifler")
    col_w(ws, {1: 18, 2: 16, 3: 14, 4: 22, 5: 22, 6: 10})

    hdr(ws, 3, ["Katman", "Seçim ★", "Alternatif", "Neden bu?", "Risk", "Değiş?"])
    tech = [
        ("Frontend", "Next.js 15", "Vue, plain HTML", "Vercel native, SSR, API routes", "Öğrenme eğrisi", "Hayır"),
        ("UI kit", "shadcn/ui", "MUI, Ant", "Hafif, özelleşir, modern", "—", "Hayır"),
        ("Backend API", "Node/TS (Fastify)", "Python FastAPI", "Tek dil, Vercel edge uyum", "—", "Hayır"),
        ("Veritabanı", "Supabase Postgres TR", "Supabase Cloud EU", "KVKK — veri TR'de", "Self-host ops", "Hayır"),
        ("Auth", "Supabase Auth + RLS", "Clerk, Auth0", "DB ile entegre, clinic_id", "—", "Hayır"),
        ("AI", "OpenAI gpt-4o-mini", "Claude, local LLM", "Maliyet/hız dengesi", "Veri dışarı", "İleride"),
        ("WhatsApp (pilot)", "Evolution self-host", "—", "Hızlı pilot", "Ban riski", "Evet→Meta"),
        ("WhatsApp (prod)", "Meta Cloud API", "Twilio", "Resmi, İYS uyumlu", "Maliyet", "F4"),
        ("Otomasyon (geçici)", "n8n Docker", "Zapier, Make", "Kontrol, TR VPS", "Teknik borç", "Evet→API"),
        ("Queue", "BullMQ + Redis", "Inngest, pg_cron", "WA/AI async", "Redis kurulum", "F3"),
        ("Hosting app", "Vercel", "VPS Caddy", "CDN, deploy kolay", "ABD edge", "Hayır"),
        ("Hosting backend", "TR VPS Docker", "Railway, Fly TR", "KVKK, 7/24", "Tek SPOF", "F5 yedek"),
        ("CI/CD", "GitHub Actions", "—", "Test + deploy otomatik", "—", "F2"),
        ("Monorepo", "nefalix-core + landing", "Tek repo", "API ve UI ayrı deploy", "—", "F2"),
    ]
    for ri, row in enumerate(tech, 4):
        for ci, v in enumerate(row, 1):
            bg = C["green_l"] if ci == 2 else (C["amber_l"] if v == "Evet→Meta" or v == "Evet→API" else None)
            cell(ws, ri, ci, v, bg=bg)


# ═══════════════════════════════════════════════════════════════
# 4 — n8n → API GEÇİŞ MATRİSİ
# ═══════════════════════════════════════════════════════════════
def sheet_migration_matrix(wb):
    ws = wb.create_sheet("4-n8n Geçiş")
    title(ws, 1, "n8n → KENDİ YAZILIM GEÇİŞ MATRİSİ", "Strangler Fig — modül modül kapat")
    col_w(ws, {1: 4, 2: 18, 3: 22, 4: 22, 5: 12, 6: 12, 7: 14, 8: 10})

    hdr(ws, 3, ["#", "Modül", "n8n'de ne var?", "API'de ne olacak?", "Başla", "Bitir", "Test", "n8n kapat"])
    mig = [
        (1, "Dashboard read", "Code node → Supabase GET", "GET /dashboard → Supabase SDK", "F0", "F2", "curl + panel", "wf-07"),
        (2, "Inbox send", "HTTP → Evolution", "POST /inbox/send", "F1", "F1", "test-all", "wf-08"),
        (3, "Inbox incoming", "Normalize + OpenAI + save", "POST /inbox/incoming + queue", "F1", "F2", "WA mesaj", "wf-06"),
        (4, "Google review", "OpenAI analiz + save", "POST /reviews/analyze", "F1", "F2", "webhook test", "wf-02"),
        (5, "NPS feedback", "Skor dallanma + WA", "POST /nps/response", "F2", "F3", "HBYS sim", "wf-01"),
        (6, "Sentinel", "OpenAI + mention save", "POST /sentinel/mention", "F3", "F3", "webhook", "wf-04"),
        (7, "Recall", "OpenAI + WA + save", "Cron /recall/run", "F3", "F4", "cron test", "wf-05"),
        (8, "eNPS", "Skor + save", "POST /enps/response", "F4", "F4", "webhook", "wf-03"),
        (9, "Web chat", "Chat trigger + OpenAI", "POST /chat/message", "F2", "F3", "site widget", "chat wf"),
    ]
    for ri, row in enumerate(mig, 4):
        for ci, v in enumerate(row, 1):
            bg = C["n8n"] if ci == 3 else (C["api"] if ci == 4 else (C["red_l"] if ci == 8 else None))
            cell(ws, ri, ci, v, bg=bg, center=(ci in (1, 5, 6, 8)))

    r = 4 + len(mig) + 2
    ws.merge_cells(f"A{r}:H{r}")
    ws[f"A{r}"] = (
        "Kural: API canlı + 2 hafta sorunsuz → n8n workflow deactivate. "
        "Geri dönüş: n8n hâlâ Docker'da, 1 tık reactivate."
    )
    ws[f"A{r}"].fill = _fill(C["amber_l"])
    ws[f"A{r}"].alignment = Alignment(wrap_text=True)


# ═══════════════════════════════════════════════════════════════
# 5 — REPO & KLASÖR YAPISI
# ═══════════════════════════════════════════════════════════════
def sheet_repo_structure(wb):
    ws = wb.create_sheet("5-Kod Yapısı")
    title(ws, 1, "REPO & KLASÖR MATRİSİ", "Kod nerede yaşar")
    col_w(ws, {1: 28, 2: 14, 3: 40, 4: 14})

    hdr(ws, 3, ["Yol / dosya", "Repo", "İçerik", "Durum"])
    structure = [
        ("nefalix-landing/", "landing", "Site + dashboard HTML/JS", "Canlı"),
        ("nefalix-landing/api/", "landing", "Vercel proxy (auth, dashboard)", "Canlı"),
        ("n8n-repo/workflows/", "n8n-repo", "Workflow JSON kaynak", "Geçici"),
        ("n8n-repo/execution/", "n8n-repo", "Deploy, test, import script", "Canlı"),
        ("n8n-repo/supabase/migrations/", "n8n-repo", "DB şema tek kaynak", "Canlı"),
        ("nefalix-core/ (yeni)", "core", "API + servisler + test", "F1'de aç"),
        ("nefalix-core/src/modules/inbox/", "core", "inbox.service.ts, prompts", "F1"),
        ("nefalix-core/src/modules/reviews/", "core", "google AI analiz", "F1"),
        ("nefalix-core/src/integrations/", "core", "evolution, openai, supabase", "F1"),
        ("nefalix-core/src/workers/", "core", "BullMQ job processor", "F3"),
        ("apps/web/ (Next.js)", "monorepo", "Dashboard React", "F2"),
        ("packages/shared/", "monorepo", "Tipler, sabitler", "F2"),
    ]
    for ri, row in enumerate(structure, 4):
        for ci, v in enumerate(row, 1):
            bg = C["api"] if "yeni" in str(v).lower() or v == "F1" else (C["done"] if v == "Canlı" else C["n8n"])
            if ci == 4:
                bg = status_color(str(v))
            cell(ws, ri, ci, v, bg=bg if ci in (1, 4) else (C["gray"] if ri % 2 == 0 else None))


# ═══════════════════════════════════════════════════════════════
# 6 — EKİP RACI
# ═══════════════════════════════════════════════════════════════
def sheet_raci(wb):
    ws = wb.create_sheet("6-Ekip RACI")
    title(ws, 1, "EKİP × SORUMLULUK (RACI)", "R=Sorumlu A=Onaylayan C=Danışılan I=Bilgilendirilen")
    col_w(ws, {1: 22, 2: 8, 3: 8, 4: 8, 5: 8, 6: 8, 7: 8})

    roles = ["Kurucu", "Dev", "Satış", "Hukuk", "Klinik"]
    hdr(ws, 3, ["Görev"] + roles)
    tasks = [
        ("Ürün önceliklendirme", "A", "C", "C", "I", "I"),
        ("API geliştirme", "I", "R", "I", "I", "I"),
        ("Dashboard UI", "C", "R", "I", "I", "C"),
        ("VPS / DevOps", "C", "R", "I", "I", "I"),
        ("KVKK metinleri", "A", "I", "I", "R", "C"),
        ("Pilot klinik eğitimi", "C", "C", "R", "I", "A"),
        ("Fiyatlandırma", "A", "I", "R", "I", "C"),
        ("n8n kapatma kararı", "A", "R", "I", "I", "I"),
        ("HBYS partner görüşme", "R", "C", "R", "I", "C"),
        ("Müşteri destek", "I", "C", "R", "I", "C"),
    ]
    raci_colors = {"R": C["teal_l"], "A": C["green_l"], "C": C["blue_l"], "I": C["gray"]}
    for ri, row in enumerate(tasks, 4):
        cell(ws, ri, 1, row[0], bold=True)
        for ci, v in enumerate(row[1:], 2):
            cell(ws, ri, ci, v, bg=raci_colors.get(v), center=True)


# ═══════════════════════════════════════════════════════════════
# 7 — ENTEGRASYON MATRİSİ
# ═══════════════════════════════════════════════════════════════
def sheet_integrations(wb):
    ws = wb.create_sheet("7-Entegrasyonlar")
    title(ws, 1, "DIŞ SİSTEM ENTEGRASYON MATRİSİ")
    col_w(ws, {1: 16, 2: 12, 3: 14, 4: 18, 5: 14, 6: 14, 7: 20})

    hdr(ws, 3, ["Sistem", "Yön", "Protokol", "Bugün", "Hedef", "Faz", "Not"])
    ints = [
        ("Evolution API", "↔ WA", "REST webhook", "Canlı pilot", "Kaldır", "F4", "Meta'ya geç"),
        ("Meta WhatsApp", "↔ Hasta", "Cloud API", "Yok", "Canlı", "F4", "İYS şart"),
        ("OpenAI", "→ AI", "REST API", "Canlı", "Canlı", "—", "Prompt TR, maskeleme"),
        ("Supabase", "↔ Veri", "PostgREST", "Canlı", "Canlı", "—", "RLS ekle"),
        ("Vercel", "→ UI", "Deploy", "Canlı", "Canlı", "—", "Sadece UI"),
        ("HBYS (Bulut Klinik)", "← Randevu", "Webhook/API", "Stub", "2-yön", "F4", "Partner gerek"),
        ("Google Business", "← Yorum", "Webhook/manual", "Simüle", "API okuma", "F3", "Yayınlama F4"),
        ("İYS", "→ İzin", "REST API", "Yok", "Middleware", "F3", "Ticari mesaj öncesi"),
        ("Site chat widget", "→ Inbox", "Webhook", "Tunnel kırık", "API", "F2", "VPS URL"),
        ("E-posta (Resend)", "→ Bildirim", "REST", "Yok", "Canlı", "F3", "Kritik alarm"),
    ]
    for ri, row in enumerate(ints, 4):
        for ci, v in enumerate(row, 1):
            cell(ws, ri, ci, v, bg=C["gray"] if ri % 2 == 0 else None)


# ═══════════════════════════════════════════════════════════════
# 8 — DEFINITION OF DONE
# ═══════════════════════════════════════════════════════════════
def sheet_dod(wb):
    ws = wb.create_sheet("8-Tamamlanma Kriteri")
    title(ws, 1, "DEFINITION OF DONE — MODÜL BAŞINA", "Bir modül 'bitti' demek için")
    col_w(ws, {1: 16, 2: 50, 3: 10})

    hdr(ws, 3, ["Modül", "Tamamlanma checklist", "Faz"])
    dod = [
        ("Inbox", "□ API test geçti □ UI bağlı □ WA gönder çalışıyor □ Supabase kayıt □ n8n kapalı □ Medident 1 hafta kullandı", "F1"),
        ("Google", "□ AI analiz API □ Onay UI □ Durum güncelleme □ n8n kapalı", "F1"),
        ("Dashboard", "□ n8n proxy yok, direkt Supabase/API □ <3sn yükleme □ Rol filtresi", "F2"),
        ("NPS", "□ HBYS webhook API □ İYS kontrol node □ Promoter/detractor akış", "F2"),
        ("Auth", "□ RLS clinic_id □ Manager sadece kendi kliniği □ Admin tümü", "F3"),
        ("Billing", "□ Fatura kaydı □ Paket limiti □ Upgrade akışı", "F3"),
        ("Production", "□ uptime %99 □ günlük yedek □ n8n container durduruldu", "F4"),
    ]
    for ri, row in enumerate(dod, 4):
        for ci, v in enumerate(row, 1):
            cell(ws, ri, ci, v, bg=C["green_l"] if ci == 3 else None)


# ═══════════════════════════════════════════════════════════════
# 9 — HAFTALIK SPRINT MATRİSİ (F1 örnek)
# ═══════════════════════════════════════════════════════════════
def sheet_sprint_matrix(wb):
    ws = wb.create_sheet("9-Sprint Matrisi")
    title(ws, 1, "SPRINT MATRİSİ — F1 ÖRNEĞİ (6 hafta)", "Satır=görev, sütun=hafta")
    col_w(ws, {1: 24, 2: 6, 3: 6, 4: 6, 5: 6, 6: 6, 7: 6, 8: 6})

    weeks = ["Hf1", "Hf2", "Hf3", "Hf4", "Hf5", "Hf6"]
    hdr(ws, 3, ["Görev", "Sahip"] + weeks)
    sprints = [
        ("nefalix-core repo bootstrap", "Dev", "■", "■", "", "", "", ""),
        ("GET /api/v1/dashboard", "Dev", "", "■", "■", "", "", ""),
        ("POST /api/v1/inbox/send", "Dev", "", "", "■", "■", "", ""),
        ("POST /api/v1/inbox/incoming + AI", "Dev", "", "", "", "■", "■", ""),
        ("Google onay UI", "Dev", "", "", "■", "■", "■", ""),
        ("Inbox thread UI", "Dev", "", "", "", "", "■", "■"),
        ("Medident UAT test", "Ops", "", "", "", "", "■", "■"),
        ("n8n wf-07/08 deactivate", "Dev", "", "", "", "", "", "■"),
    ]
    for ri, row in enumerate(sprints, 4):
        cell(ws, ri, 1, row[0])
        cell(ws, ri, 2, row[1], center=True)
        for wi, mark in enumerate(row[2:], 3):
            c = cell(ws, ri, wi, mark, center=True)
            if mark == "■":
                c.fill = _fill(C["teal_l"])
                c.font = _f(bold=True, color=C["teal_d"])


# ═══════════════════════════════════════════════════════════════
# 10 — BUILD vs BUY
# ═══════════════════════════════════════════════════════════════
def sheet_build_buy(wb):
    ws = wb.create_sheet("10-Build vs Buy")
    title(ws, 1, "YAP vs AL MATRİSİ")
    col_w(ws, {1: 20, 2: 10, 3: 22, 4: 30, 5: 14})

    hdr(ws, 3, ["Bileşen", "Karar", "Seçenek", "Gerekçe", "Maliyet/ay"])
    bb = [
        ("Inbox + AI mantığı", "YAP", "Kendi API", "Fark yaratan çekirdek IP", "—"),
        ("Veritabanı", "YAP/HOST", "Supabase self-host", "KVKK kontrol", "170 ₺"),
        ("Auth", "AL", "Supabase Auth", "Güvenlik riski yüksek", "0"),
        ("Ödeme/fatura", "AL", "iyzico / Parasut", "PCI yükü", "~500 ₺"),
        ("WhatsApp (prod)", "AL", "Meta Cloud API", "Resmi kanal şart", "~2000 ₺"),
        ("E-posta", "AL", "Resend", "Deliverability", "0-200 ₺"),
        ("Monitoring", "AL", "UptimeRobot + Sentry", "Hızlı kurulum", "0-400 ₺"),
        ("n8n (geçici)", "AL/self", "Docker n8n", "Prototip hızı", "0"),
        ("HBYS", "PARTNER", "Bulut Klinik API", "Yeniden icat etme", "Teklif"),
        ("CRM satış", "AL", "Notion / HubSpot free", "Erken aşama", "0"),
    ]
    for ri, row in enumerate(bb, 4):
        for ci, v in enumerate(row, 1):
            bg = C["green_l"] if v == "YAP" else (C["blue_l"] if v == "AL" else (C["purple_l"] if v == "PARTNER" else C["amber_l"]))
            if ci == 2:
                cell(ws, ri, ci, v, bg=bg, center=True, bold=True)
            else:
                cell(ws, ri, ci, v, bg=C["gray"] if ri % 2 == 0 else None)


def main():
    wb = Workbook()
    sheet_master(wb)
    sheet_module_matrix(wb)
    sheet_phase_matrix(wb)
    sheet_tech_matrix(wb)
    sheet_migration_matrix(wb)
    sheet_repo_structure(wb)
    sheet_raci(wb)
    sheet_integrations(wb)
    sheet_dod(wb)
    sheet_sprint_matrix(wb)
    sheet_build_buy(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"✓ SaaS matrisi ({len(wb.sheetnames)} sayfa): {OUT}")


if __name__ == "__main__":
    main()
