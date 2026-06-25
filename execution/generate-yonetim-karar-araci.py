#!/usr/bin/env python3
"""
NefalixAI — Yönetim Karar Aracı + Portföy (referans format).
Örnek: Kanal Çakışması Karar Aracı + İstanbul ICP Portföy stili.
"""
from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent.parent / "docs" / "yonetici-sunum" / "NefalixAI_Yonetim_ve_Buyume_Karar_Araci.xlsx"
ICP_SRC = Path("/Users/enesceylan/Downloads/NefalixAI Istanbul ICP Portfoy.xlsx")
TODAY = date(2026, 6, 18)

NAVY = "0F2040"
NAVY2 = "1A2E45"
TEAL = "0D9488"
WHITE = "FFFFFF"
INPUT_BG = "FFF6E6"
INPUT_FG = "0000FF"
HDR_BG = "0F2040"
SEC_BG = "E8F4F8"
GREEN_L = "E8F8EE"
AMBER_L = "FFF6E6"
RED_L = "FDE8E8"
GRAY = "F5F7FA"
BORDER_C = "D0D7DE"


def side():
    return Side(style="thin", color=BORDER_C)


def border():
    return Border(left=side(), right=side(), top=side(), bottom=side())


def fill(h):
    return PatternFill("solid", fgColor=h)


def fnt(bold=False, size=10, color=NAVY2, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


def merge_title(ws, row, text, subtitle=None, cols=8):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=cols)
    c = ws.cell(row, 2, text)
    c.font = fnt(bold=True, size=16, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28
    if subtitle:
        r2 = row + 1
        ws.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=cols)
        ws.cell(r2, 2, subtitle).font = fnt(size=10, color=NAVY2, italic=True)
        ws.cell(r2, 2).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r2].height = 32


def section(ws, row, text, cols=8):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=cols)
    c = ws.cell(row, 2, text)
    c.font = fnt(bold=True, size=11, color=NAVY)
    c.fill = fill(SEC_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def col_w(ws, widths: dict):
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w


def hdr_row(ws, row, headers, start_col=2):
    for i, h in enumerate(headers, start_col):
        c = ws.cell(row, i, h)
        c.font = fnt(bold=True, size=10, color=WHITE)
        c.fill = fill(HDR_BG)
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 36


def write_cell(ws, r, c, val, input_cell=False, center=False, bold=False, bg=None):
    x = ws.cell(r, c, val)
    x.border = border()
    x.alignment = Alignment(wrap_text=True, vertical="top",
                            horizontal="center" if center else "left")
    if input_cell:
        x.fill = fill(INPUT_BG)
        x.font = fnt(bold=True, color=INPUT_FG)
    elif bg:
        x.fill = fill(bg)
    if bold:
        x.font = fnt(bold=True)
    return x


def dropdown(ws, cell_ref, options: str):
    dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
    dv.error = "Listeden seçin"
    dv.errorTitle = "Geçersiz seçim"
    ws.add_data_validation(dv)
    dv.add(cell_ref)


# ═══════════════════════════════════════════════════════════════
def sheet_metodoloji(wb):
    ws = wb.active
    ws.title = "Metodoloji ve Kullanım"
    ws.sheet_view.showGridLines = False
    col_w(ws, {1: 3, 2: 28, 3: 55, 4: 3})

    merge_title(ws, 2, "NEFALIXAI — YÖNETİM VE BÜYÜME KARAR ARACI",
                "Yönetici sunumu, satış teşhisi, maliyet-gelir ve İstanbul hedef portföyü — tek dosyada", 3)

    blocks = [
        ("1. BU DOSYA NE İŞE YARAR?", [
            ("Amaç", "Yöneticilerin teknik bilgi olmadan karar vermesi: hangi paketi satacağız, ne zaman, kime, ne kadar yatırım."),
            ("Kapsam", "Pilot (Medident) durumu, 12 aylık yol haritası, maliyet-gelir, riskler, İstanbul diş kliniği hedef listesi."),
        ]),
        ("2. SAYFA REHBERİ", [
            ("Metodoloji ve Kullanım", "Bu sayfa — nasıl okunur."),
            ("Özet Panel", "Canlı sayılar — diğer sayfalardan otomatik toplanır."),
            ("Klinik Satış Teşhis Aracı", "3 soru sor → hangi paket, hangi öncelik çıksın (açılır liste)."),
            ("Paket ve Senaryo Referansı", "Teşhis sonucunun detaylı açıklaması."),
            ("Yol Haritası Referansı", "Ay ay ne yapılacak — teknik jargon yok."),
            ("Maliyet ve Gelir", "Aylık giderler ve 4 gelir senaryosu."),
            ("Risk Referansı", "Ne ters gidebilir, ne yapıyoruz."),
            ("Pilot — Medident", "İlk müşteri skor kartı."),
            ("1. Diş Klinikleri", "İstanbul hedef listesi — satış önceliklendirme."),
            ("Toplantı Onay Listesi", "Yönetim kurulu EVET/HAYIR."),
        ]),
        ("3. NASIL KULLANILIR (5 DAKİKA)", [
            ("Adım 1", "Özet Panel'i aç — genel tabloyu gör."),
            ("Adım 2", "Klinik Satış Teşhis Aracı'nda 3 soruyu doldur (açılır listeler)."),
            ("Adım 3", "Çıkan paket önerisini Paket ve Senaryo Referansı'nda oku."),
            ("Adım 4", "Toplantı Onay Listesi'nde kararları işaretle."),
        ]),
        ("4. TEMEL İLKELER (YÖNETİCİ DİLİ)", [
            ("Veri Türkiye'de", "Hasta bilgisi yurt dışına gitmemeli — sunucu İstanbul."),
            ("Önce çalışsın", "Güzel arayüz sonra — önce Medident günlük kullanmalı."),
            ("Yıl 1 yatırım", "İlk yıl kâr beklenmez; 8–10 klinikte anlamlı gelir."),
            ("Fiyat önerisi", "Orta paket 7.990 ₺/ay + kurulum 14.900 ₺."),
        ]),
    ]
    r = 5
    for title, items in blocks:
        section(ws, r, title, 3)
        r += 1
        for k, v in items:
            write_cell(ws, r, 2, k, bold=True)
            write_cell(ws, r, 3, v)
            ws.row_dimensions[r].height = 28 if len(str(v)) > 60 else 20
            r += 1
        r += 1


def sheet_ozet_panel(wb):
    ws = wb.create_sheet("Özet Panel")
    ws.sheet_view.showGridLines = False
    col_w(ws, {1: 3, 2: 22, 3: 14, 4: 16, 5: 3, 6: 22, 7: 14})

    merge_title(ws, 2, "NEFALIXAI — YÖNETİM ÖZET PANELİ",
                "Tüm sayfalardan canlı toplanan sayılar — sunumda ilk açılacak sayfa", 7)

    section(ws, 5, "PORTFÖY ÖZETİ", 7)
    hdr_row(ws, 6, ["Gösterge", "Değer", "Açıklama"], 2)
    rows = [
        ("Hedef diş kliniği (listedeki)", "=COUNTA('1. Diş Klinikleri'!B7:B40)", "İstanbul portföy"),
        ("Yüksek öncelik (skor ≥4)", "=COUNTIF('1. Diş Klinikleri'!I7:I40,\">=4\")", "Önce aranacaklar"),
        ("Pilot klinik", "Medident İstanbul", "Aktif pilot"),
        ("Önerilen aylık paket", "7.990 ₺", "Profesyonel"),
        ("12 ay hedef klinik", "8", "Gerçekçi senaryo"),
        ("12 ay hedef aylık gelir", "63.920 ₺", "8 × 7.990"),
        ("Pilot aylık maliyet (yazılımcı hariç)", "~6.000 ₺", "Maliyet sayfası"),
        ("Satışa hazır aylık maliyet", "~82.000 ₺", "Yazılımcı dahil"),
    ]
    for i, (a, b, c) in enumerate(rows, 7):
        write_cell(ws, i, 2, a, bold=True)
        write_cell(ws, i, 3, b, bg=GREEN_L if i % 2 == 0 else GRAY, center=True)
        write_cell(ws, i, 4, c)

    section(ws, 16, "GELİR SENARYOLARI (YIL 1)", 7)
    hdr_row(ws, 17, ["Senaryo", "Klinik", "Yıllık gelir (₺)"], 2)
    scenarios = [
        ("Kötü", 3, 216000),
        ("Gerçekçi ★", 8, 718000),
        ("İyi", 15, 1440000),
        ("Çok iyi", 25, 2550000),
    ]
    for i, (s, n, g) in enumerate(scenarios, 18):
        write_cell(ws, i, 2, s, bold=True, bg=AMBER_L if "★" in s else None)
        write_cell(ws, i, 3, n, center=True)
        write_cell(ws, i, 4, g, center=True)

    section(ws, 23, "SUNUCU / PİLOT DURUMU", 7)
    status = [
        ("VPS durumu", "⏳ Kimlik doğrulama — destek açık"),
        ("Panel adresi", "nefalixai.com/dashboard"),
        ("Sonraki adım", "Sunucu açılınca Medident eğitimi"),
    ]
    for i, (a, b) in enumerate(status, 24):
        write_cell(ws, i, 2, a, bold=True)
        write_cell(ws, i, 3, b)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)


def sheet_teshis(wb):
    ws = wb.create_sheet("Klinik Satış Teşhis Aracı")
    ws.sheet_view.showGridLines = False
    col_w(ws, {1: 3, 2: 32, 3: 36, 4: 28})

    merge_title(ws, 2, "NEFALIXAI — KLİNİK SATIŞ TEŞHİS ARACI",
                "Satış görüşmesinde 3 soruyu sorun, C sütunundan seçin — sonuç otomatik gelir", 4)

    section(ws, 5, "ADIM 1 — KLİNİK BİLGİLERİ", 4)
    write_cell(ws, 6, 2, "Klinik Adı", bold=True)
    write_cell(ws, 6, 3, "", input_cell=True)
    write_cell(ws, 7, 2, "Görüşme Tarihi", bold=True)
    write_cell(ws, 7, 3, TODAY.strftime("%d.%m.%Y"), input_cell=True)

    section(ws, 9, "ADIM 2 — ÜÇ SORUYU SORUN, CEVABI SEÇİN", 4)
    qs = [
        ("SORU 1 — Klinik ölçeği", '"Kaç hekiminiz var, tek şube mi?"', "Tek hekim,Tek şube orta (2-5 hekim),Büyük / çok şube"),
        ("SORU 2 — WhatsApp yoğunluğu", '"Günde kaç WhatsApp mesajı alıyorsunuz?"', "Az (günde 10 altı),Orta (10-30),Yoğun (30 üzeri)"),
        ("SORU 3 — En acil ihtiyaç", '"En çok neyi çözmek istiyorsunuz?"', "Mesaj kaçırma,Google yorum ve itibar,Hepsi — tam paket"),
    ]
    for i, (q, hint, opts) in enumerate(qs, 10):
        write_cell(ws, i, 2, q, bold=True)
        write_cell(ws, i, 3, "", input_cell=True)
        dropdown(ws, f"C{i}", opts)
        write_cell(ws, i, 4, hint, bg=GRAY)

    section(ws, 14, "ADIM 3 — SONUÇ", 4)
    write_cell(ws, 15, 2, "Önerilen paket ve yaklaşım", bold=True)
    # Lookup based on Q1+Q3 primarily
    formula = (
        '=IF(OR(C10="",C11="",C12=""),"Önce 3 soruyu da cevaplayın",'
        'IF(C12="Hepsi — tam paket","SENARYO P — Profesyonel paket (7.990 ₺/ay) + kurulum 14.900 ₺",'
        'IF(C12="Google yorum ve itibar","SENARYO G — Profesyonel (yorum modülü öncelikli)",'
        'IF(AND(C10="Tek hekim",C11="Az (günde 10 altı)"),"SENARYO S — Başlangıç paketi (4.990 ₺/ay)",'
        'IF(C10="Büyük / çok şube","SENARYO K — Kurumsal paket (14.990 ₺/ay) — özel teklif hazırlayın",'
        '"SENARYO P — Profesyonel paket (7.990 ₺/ay)")))))'
    )
    c = write_cell(ws, 15, 3, formula)
    c.font = fnt(bold=True, size=11, color=NAVY)
    c.fill = fill(GREEN_L)
    ws.row_dimensions[15].height = 48
    write_cell(ws, 15, 4, "Detay için 'Paket ve Senaryo Referansı' sayfasına bakın", bg=GRAY)

    write_cell(ws, 16, 2, "Tahmini aylık gelir (bizim)", bold=True)
    write_cell(ws, 16, 3,
               '=IF(C15="","",IF(ISNUMBER(SEARCH("4.990",C15)),4990,IF(ISNUMBER(SEARCH("14.990",C15)),14990,7990)))',
               center=True, bg=AMBER_L)
    write_cell(ws, 17, 2, "Kurulum ücreti önerisi", bold=True)
    write_cell(ws, 17, 3,
               '=IF(C15="","",IF(ISNUMBER(SEARCH("4.990",C15)),9900,IF(ISNUMBER(SEARCH("14.990",C15)),29900,14900)))',
               center=True, bg=AMBER_L)

    section(ws, 19, "NASIL KULLANILIR", 4)
    notes = (
        "1) Klinik adını ve tarihi girin.\n"
        "2) Görüşmede 3 soruyu sorun — cevapları C sütunundan seçin.\n"
        "3) ADIM 3'te paket önerisi çıkar — referans sayfasında detayı okuyun.\n"
        "4) Toplantı Onay Listesi'ne geçin."
    )
    ws.merge_cells("B20:D22")
    ws["B20"] = notes
    ws["B20"].alignment = Alignment(wrap_text=True)
    ws["B20"].font = fnt(size=10)


def sheet_paket_ref(wb):
    ws = wb.create_sheet("Paket ve Senaryo Referansı")
    col_w(ws, {1: 3, 2: 14, 3: 22, 4: 28, 5: 18, 6: 12, 7: 28, 8: 14})

    merge_title(ws, 2, "PAKET VE SENARYO — DETAYLI REFERANS",
                "Klinik Satış Teşhis Aracı'nda çıkan senaryonun tam açıklaması", 8)

    hdr_row(ws, 5, ["Kod", "Klinik profili", "Önerilen paket", "Satış cümlesi", "Aylık (₺)", "Klinikle anlaşma cümlesi", "İlk modül"])

    data = [
        ("S", "Tek hekim · az mesaj", "Başlangıç 4.990 ₺", "Küçük ama düzenli iletişim", 4990,
         '"WhatsApp ve site sohbetini tek yerden yönetin — ayda 5 bin ₺\'den başlar."', "Inbox + chat"),
        ("P", "Orta klinik · orta/yüksek mesaj", "Profesyonel 7.990 ₺", "Ana hedef paket", 7990,
         '"Bir implant randevusu programı 2 yıl öder. Mesaj kaçırmaz, yorumları yönetirsiniz."', "Inbox + Google + NPS"),
        ("G", "Yorum/itibar öncelikli", "Profesyonel 7.990 ₺", "Google odaklı satış", 7990,
         '"Google puanınızı yükseltmek için yapay zeka taslak + onay paneli."', "Google yorum"),
        ("K", "Büyük / çok şube", "Kurumsal 14.990 ₺", "Kurumsal teklif", 14990,
         '"Tüm şubelerinizi tek panelden yönetin — özel eğitim ve destek."', "Çoklu klinik + SLA"),
        ("H", "Hiç dijital takip yok", "Profesyonel 7.990 ₺", "En kolay satış", 7990,
         '"Şu an hiç takip etmiyorsunuz — biz sıfırdan kurarız, siz sadece onaylarsınız."', "Tam kurulum"),
    ]
    for i, row in enumerate(data, 6):
        for j, v in enumerate(row, 2):
            write_cell(ws, i, j, v, bg=GRAY if i % 2 else None)
        if row[0] == "P":
            for j in range(2, 9):
                ws.cell(i, j).fill = fill(GREEN_L)

    r = 12
    section(ws, r, "PAKET KARŞILAŞTIRMA", 8)
    hdr_row(ws, r + 1, ["Özellik", "Başlangıç", "Profesyonel ★", "Kurumsal"], 2)
    feats = [
        ("WhatsApp gelen kutusu", "✓", "✓", "✓"),
        ("Yapay zeka taslak yanıt", "✓", "✓", "✓"),
        ("Google yorum analizi", "—", "✓", "✓"),
        ("Hasta memnuniyeti (NPS)", "—", "✓", "✓"),
        ("İtibar takibi", "—", "✓", "✓"),
        ("Geri çağırma", "—", "✓", "✓"),
        ("Çoklu şube", "—", "—", "✓"),
    ]
    for i, row in enumerate(feats, r + 2):
        for j, v in enumerate(row, 2):
            write_cell(ws, i, j, v, center=(j > 2))


def sheet_yol_ref(wb):
    ws = wb.create_sheet("Yol Haritası Referansı")
    col_w(ws, {1: 3, 2: 14, 3: 28, 4: 28, 5: 12, 6: 28})
    merge_title(ws, 2, "YOL HARİTASI — DETAYLI REFERANS", "Yöneticiler için — teknik terim yok", 6)
    hdr_row(ws, 5, ["Dönem", "Klinik ne görür?", "Biz ne yaparız?", "Harcama", "Bitti sayılır eğer…"])
    rows = [
        ("Haz 2026", "Panel açılır, veriler gelir", "Sunucu + program çalıştırma", "Düşük", "Medident panele girer"),
        ("Tem 2026", "WhatsApp yanıtı panelden", "Gelen kutusu iyileştirme", "Orta", "Haftada 5 gün kullanım"),
        ("Ağu–Eyl", "Google yorum onayı tek tık", "Onay ekranı", "Orta", "Yorumların %50'si panelden"),
        ("Eki–Ara", "İlk ücretli müşteriler", "Satış, sözleşme, eğitim", "Yüksek", "5 klinik ödüyor"),
        ("2027 H1", "Randevu sistemi bağlantısı", "Klinik yazılımları entegrasyonu", "Yüksek", "10+ klinik"),
    ]
    for i, row in enumerate(rows, 6):
        for j, v in enumerate(row, 2):
            write_cell(ws, i, j, v, bg=GRAY if i % 2 else None)


def sheet_maliyet_gelir(wb):
    ws = wb.create_sheet("Maliyet ve Gelir")
    col_w(ws, {1: 3, 2: 24, 3: 32, 4: 14, 5: 12})
    merge_title(ws, 2, "MALİYET VE GELİR TABLOSU", "Tahmini rakamlar — güncel teklif alın", 5)

    section(ws, 5, "AYLIK GİDERLER", 5)
    hdr_row(ws, 6, ["Kalem", "Ne için?", "Ayda (₺)", "Zorunlu?"])
    costs = [
        ("Türkiye sunucu", "Program 7/24, veri TR'de", 170, "Evet"),
        ("Site barındırma", "nefalixai.com", 760, "Evet"),
        ("Yapay zeka", "Taslak mesaj / yorum", 1500, "Evet"),
        ("WhatsApp resmi hat", "Yasal ticari mesaj", 2000, "Satışta"),
        ("Hukuk / KVKK", "Sözleşme", 2500, "Evet"),
        ("Yazılımcı", "Geliştirme", 45000, "Önerilen"),
        ("", "PİLOT TOPLAM (yazılımcı hariç)", "~6.000", ""),
        ("", "SATIŞA HAZIR (dahil)", "~82.000", ""),
    ]
    for i, row in enumerate(costs, 7):
        for j, v in enumerate(row, 2):
            write_cell(ws, i, j, v, bold=(row[0] == ""))

    section(ws, 17, "YIL 1 GELİR SENARYOLARI", 5)
    hdr_row(ws, 18, ["Senaryo", "Klinik sayısı", "Ort. aylık", "Yıllık gelir (₺)", "Not"])
    for i, row in enumerate([
        ("Kötü", 3, 5990, 216000, "Satış yavaş"),
        ("Gerçekçi ★", 8, 7490, 718000, "Medident referansı"),
        ("İyi", 15, 7990, 1440000, "İyi satış"),
        ("Çok iyi", 25, 8490, 2550000, "Hızlı büyüme"),
    ], 19):
        for j, v in enumerate(row, 2):
            write_cell(ws, i, j, v, bg=GREEN_L if "★" in str(row[0]) else None)


def sheet_risk_ref(wb):
    ws = wb.create_sheet("Risk Referansı")
    col_w(ws, {1: 3, 2: 20, 3: 28, 4: 28, 5: 12})
    merge_title(ws, 2, "RİSK REFERANSI", "Olasılık ve azaltma planı", 5)
    hdr_row(ws, 5, ["Risk", "Ne olabilir?", "Ne yapıyoruz?", "Önem"])
    risks = [
        ("Sunucu kapalı", "Panel çalışmaz", "TR sunucu + yedekleme", "Yüksek"),
        ("WhatsApp kapanır", "Mesaj gitmez", "Resmi iş ortağı hattı", "Yüksek"),
        ("Yasa ihlali", "Ceza", "Avukat + İYS", "Çok yüksek"),
        ("Satış olmaz", "Gelir yok", "Medident hikâyesi", "Orta"),
    ]
    for i, row in enumerate(risks, 6):
        for j, v in enumerate(row, 2):
            c = write_cell(ws, i, j, v)
            if j == 5 and v == "Çok yüksek":
                c.fill = fill(RED_L)


def sheet_medident(wb):
    ws = wb.create_sheet("Pilot — Medident")
    col_w(ws, {1: 3, 2: 22, 3: 14, 4: 14, 5: 28})
    merge_title(ws, 2, "PİLOT KLİNİK — MEDIDENT İSTANBUL", "Abdulkadir Yaşar · medidentistanbul.com", 5)
    hdr_row(ws, 5, ["Gösterge", "Hedef", "Şu an", "Not"])
    rows = [
        ("Panel erişimi", "%99", "Sunucu askı", "Destek açık"),
        ("Haftalık kullanım", "5+ gün", "—", "Sunucu sonrası"),
        ("Inbox panelden", "%80", "—", ""),
        ("Google onay", "%50", "—", ""),
        ("WhatsApp bağlı", "Evet", "QR gerekli", ""),
        ("Referans hazır", "Evet", "Hayır", "Case study"),
    ]
    for i, row in enumerate(rows, 6):
        for j, v in enumerate(row, 2):
            write_cell(ws, i, j, v)


def sheet_onay(wb):
    ws = wb.create_sheet("Toplantı Onay Listesi")
    col_w(ws, {1: 3, 2: 4, 3: 36, 4: 16, 5: 16, 6: 20})
    merge_title(ws, 2, "YÖNETİM KURULU ONAY LİSTESİ", "Toplantı sonunda doldurun", 6)
    hdr_row(ws, 5, ["#", "Karar", "Tutar", "ONAY (EVET/HAYIR)", "Not"])
    rows = [
        (1, "Yıllık bütçe — yatırım yılı", "~2.100.000 ₺"),
        (2, "Hukuk / KVKK paketi", "25.000 ₺"),
        (3, "Yarı zamanlı yazılımcı", "45.000 ₺/ay"),
        (4, "Fiyat: Profesyonel 7.990 ₺/ay", "—"),
        (5, "Medident referans izni", "—"),
        (6, "Hedef: 5 ücretli klinik (Aralık)", "—"),
    ]
    for i, (n, k, t) in enumerate(rows, 6):
        write_cell(ws, i, 2, n, center=True)
        write_cell(ws, i, 3, k)
        write_cell(ws, i, 4, t, center=True)
        write_cell(ws, i, 5, "", input_cell=True)
        dropdown(ws, f"E{i}", "EVET,HAYIR,KARARSIZ")
        write_cell(ws, i, 6, "")


def copy_dental_sheet(wb):
    """ICP dosyasından diş klinikleri + Medident pilot satırı."""
    ws = wb.create_sheet("1. Diş Klinikleri")
    col_w(ws, {1: 3, 2: 32, 3: 16, 4: 28, 5: 18, 6: 22, 7: 18, 8: 16, 9: 12, 10: 36})

    merge_title(ws, 2, "1. DİŞ KLİNİKLERİ — İSTANBUL HEDEF PORTFÖY",
                "Ana satış listesi · Uygunluk skoru 1-5 · Medident pilot en üstte", 10)

    headers = ["", "İşletme Adı", "Semt / İlçe", "Adres", "Telefon", "Web Sitesi",
               "Tahmini Ölçek", "Uluslararası Hasta", "Uygunluk Skoru (1-5)", "Not / Kaynak"]
    hdr_row(ws, 5, headers[1:], 2)
    ws.cell(5, 2, "")

    # Pilot row
    pilot = [
        "★ PİLOT", "Medident İstanbul", "Üsküdar", "Acıbadem Cd. 195F, Üsküdar", "+90 549 119 08 19",
        "medidentistanbul.com", "Orta klinik", "Evet", 5, "Aktif pilot — Abdulkadir Yaşar",
    ]
    for j, v in enumerate(pilot, 2):
        c = write_cell(ws, 6, j, v, bold=(j == 2))
        if j == 9:
            c.fill = fill(GREEN_L)

    if ICP_SRC.exists():
        src = load_workbook(ICP_SRC, data_only=True)["1. Diş Klinikleri"]
        r_out = 7
        for r_in in range(6, 40):
            name = src.cell(r_in, 2).value
            if not name or "Medident" in str(name):
                continue
            for c_in in range(2, 11):
                write_cell(ws, r_out, c_in, src.cell(r_in, c_in).value, bg=GRAY if r_out % 2 else None)
            r_out += 1
    else:
        write_cell(ws, 7, 2, "(ICP kaynak dosyası bulunamadı — İstanbul ICP Portföy xlsx yükleyin)")

    ws.cell(42, 2, "Not: Skor ≥4 olanlar önce aranır. Pilot tamamlandıktan sonra referans olarak kullanın.").font = fnt(italic=True, size=9)


def reorder_sheets(wb):
    order = [
        "Metodoloji ve Kullanım", "Özet Panel", "Klinik Satış Teşhis Aracı",
        "Paket ve Senaryo Referansı", "Yol Haritası Referansı", "Maliyet ve Gelir",
        "Risk Referansı", "Pilot — Medident", "1. Diş Klinikleri", "Toplantı Onay Listesi",
    ]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))


def main():
    wb = Workbook()
    sheet_metodoloji(wb)
    sheet_ozet_panel(wb)
    sheet_teshis(wb)
    sheet_paket_ref(wb)
    sheet_yol_ref(wb)
    sheet_maliyet_gelir(wb)
    sheet_risk_ref(wb)
    sheet_medident(wb)
    copy_dental_sheet(wb)
    sheet_onay(wb)
    reorder_sheets(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"✓ Karar aracı ({len(wb.sheetnames)} sayfa): {OUT}")
    print("  Google Drive → Yükle → Google E-Tablolar ile aç")


if __name__ == "__main__":
    main()
