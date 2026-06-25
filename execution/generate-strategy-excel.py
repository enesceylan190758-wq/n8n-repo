#!/usr/bin/env python3
"""
Nefalix — Gelişmiş yönetici strateji raporu (Excel v2).
Formüllü finansal model, grafikler, duyarlılık analizi, uyumluluk, sunum sayfası.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent.parent / "docs" / "Nefalix_Strateji_Yol_Haritasi_2026.xlsx"
TODAY = date(2026, 6, 18)

# ── Palet ────────────────────────────────────────────────────
C = {
    "navy": "0F172A",
    "teal": "0D9488",
    "teal_d": "0F766E",
    "teal_l": "CCFBF1",
    "slate": "64748B",
    "white": "FFFFFF",
    "amber": "F59E0B",
    "amber_l": "FEF3C7",
    "red": "DC2626",
    "red_l": "FEE2E2",
    "green": "16A34A",
    "green_l": "DCFCE7",
    "blue_l": "DBEAFE",
    "purple_l": "EDE9FE",
    "gray": "F8FAFC",
    "border": "E2E8F0",
    "gold": "CA8A04",
}


def _side():
    return Side(style="thin", color=C["border"])


def _border():
    return Border(left=_side(), right=_side(), top=_side(), bottom=_side())


def _fill(h: str):
    return PatternFill("solid", fgColor=h)


def _f(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


def col_w(ws, widths: dict[int, float]):
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def title_block(ws, row: int, title: str, sub: str = "", merge_end: str = "H"):
    ws.merge_cells(f"A{row}:{merge_end}{row}")
    ws[f"A{row}"] = title
    ws[f"A{row}"].font = _f(bold=True, size=20, color=C["navy"])
    if sub:
        r2 = row + 1
        ws.merge_cells(f"A{r2}:{merge_end}{r2}")
        ws[f"A{r2}"] = sub
        ws[f"A{r2}"].font = _f(size=11, color=C["slate"], italic=True)


def header_row(ws, row: int, headers: list[str], bg: str = C["teal"]):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = _f(bold=True, size=10, color=C["white"])
        cell.fill = _fill(bg)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_rows(ws, start: int, headers: list[str], rows: list, widths: dict | None = None, zebra=True):
    header_row(ws, start, headers)
    for ri, row in enumerate(rows, start + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if zebra and ri % 2 == 0:
                cell.fill = _fill(C["gray"])
    if widths:
        col_w(ws, widths)
    return start + len(rows) + 2


def kpi_card(ws, cell: str, label: str, value, sub: str = "", bg: str = C["navy"]):
    """2x2 merged KPI kartı — cell sol üst köşe."""
    c0 = cell[0]
    r0 = int(cell[1:])
    ws.merge_cells(f"{c0}{r0}:{chr(ord(c0)+1)}{r0}")
    ws.merge_cells(f"{c0}{r0+1}:{chr(ord(c0)+1)}{r0+1}")
    top = ws[cell]
    top.value = label
    top.font = _f(size=9, color=C["teal_l"])
    top.fill = _fill(bg)
    top.alignment = Alignment(horizontal="center")
    val_cell = f"{c0}{r0+1}"
    ws[val_cell] = value
    ws[val_cell].font = _f(bold=True, size=16, color=C["white"])
    ws[val_cell].fill = _fill(bg)
    ws[val_cell].alignment = Alignment(horizontal="center")
    if sub:
        ws.merge_cells(f"{c0}{r0+2}:{chr(ord(c0)+1)}{r0+2}")
        ws[f"{c0}{r0+2}"] = sub
        ws[f"{c0}{r0+2}"].font = _f(size=8, color=C["teal_l"], italic=True)
        ws[f"{c0}{r0+2}"].fill = _fill(bg)
        ws[f"{c0}{r0+2}"].alignment = Alignment(horizontal="center")


# ═══════════════════════════════════════════════════════════════
# SAYFA 0 — YÖNETİCİ DASHBOARD
# ═══════════════════════════════════════════════════════════════
def sheet_dashboard(wb: Workbook):
    ws = wb.active
    ws.title = "0-Dashboard"
    ws.sheet_view.showGridLines = False
    col_w(ws, {1: 3, 2: 16, 3: 16, 4: 16, 5: 16, 6: 3, 7: 16, 8: 16})

    title_block(ws, 1, "NEFALIX AI — YÖNETİCİ KONTROL PANELİ",
                f"{TODAY.strftime('%d %B %Y')} | Medident İstanbul Pilot | Gizli — Yönetim Kurulu")

    # KPI kartları — formüller Finans sayfasına bağlı
    kpi_card(ws, "B3", "HEDEF MRR (12. ay)", "=Finans!N35", "Formül: finans modeli", C["teal_d"])
    kpi_card(ws, "D3", "HEDEF KLİNİK", "=Finans!N33", "Aktif ödeyen", C["navy"])
    kpi_card(ws, "F3", "BAŞABAŞ AYI", "=Finans!N37", "MRR ≥ toplam gider", C["teal_d"])
    kpi_card(ws, "H3", "YIL 1 NET", "=Finans!N39", "Kümülatif nakit", C["navy"])

    kpi_card(ws, "B7", "ÖNERİLEN PAKET", "7.990 ₺/ay", "Profesyonel", C["teal"])
    kpi_card(ws, "D7", "BRÜT MARJ", "%82", "Klinik başı", C["teal"])
    kpi_card(ws, "F7", "LTV / CAC", "12x", "Sağlıklı >3x", C["teal"])
    kpi_card(ws, "H7", "PİLOT MALİYET", "~6.000 ₺/ay", "Meta API hariç", C["amber"])

    r = 11
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "TEK CÜMLELİK HİKÂYE (yönetim kuruluna okunur)"
    ws[f"B{r}"].font = _f(bold=True, size=12, color=C["navy"])
    r += 1
    ws.merge_cells(f"B{r}:H{r+2}")
    ws[f"B{r}"] = (
        "Nefalix, diş kliniklerinin WhatsApp mesajlarını kaçırmamasını, memnun hastaları Google'a yönlendirmesini "
        "ve şikâyetleri erken yakalamasını sağlayan Türkiye sunuculu bir yazılımdır. Medident'te pilot kuruldu; "
        "sunucu kimlik doğrulaması nedeniyle geçici duraklama var. Hedef: 12 ayda 8 ödeyen klinik, "
        "aylık ~64.000 ₺ tekrarlayan gelir. Yıl 1 yatırım yılıdır (~2,1M ₺); yıl 2'de kâra geçiş mümkün."
    )
    ws[f"B{r}"].alignment = Alignment(wrap_text=True)
    ws[f"B{r}"].fill = _fill(C["teal_l"])
    ws.row_dimensions[r].height = 72

    r += 4
    write_rows(ws, r, ["Öncelik", "Ne?", "Ne zaman?", "Kim?", "Bütçe"],
               [
                   ["P0", "VPS askıyı kaldır, pilotu başlat", "Haz 2026", "Teknik", "0 ₺"],
                   ["P0", "KVKK + İYS hukuki altyapı", "Tem 2026", "Hukuk", "31.000 ₺"],
                   ["P1", "Inbox + Google onay ekranı", "Ağu 2026", "Geliştirme", "Dahili"],
                   ["P1", "Medident case study + 2. klinik", "Eyl 2026", "Satış", "3.000 ₺"],
                   ["P2", "WhatsApp resmi API geçişi", "Kas 2026", "Teknik", "~24.000 ₺/yıl"],
               ], {1: 8, 2: 32, 3: 12, 4: 12, 5: 14})

    r = 22
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "SAYFA REHBERİ — bu dosyada nerede ne var?"
    ws[f"B{r}"].font = _f(bold=True, size=12, color=C["navy"])
    r += 1
    guide = [
        ["0-Dashboard", "Bu sayfa — özet KPI"],
        ["A-Girdiler", "Varsayımları değiştirin → finans otomatik güncellenir"],
        ["Finans", "24 aylık nakit akışı + grafikler"],
        ["Duyarlılık", "Fiyat / klinik sayısı what-if"],
        ["API Maliyet", "Klinik hacmine göre API hesabı"],
        ["Yol Haritası", "Fazlar + Gantt"],
        ["Fiyatlandırma", "Paketler + ROI"],
        ["Uyumluluk", "KVKK / İYS checklist"],
        ["Uzman Görüşmeleri", "5 disiplin Q&A"],
        ["Sunum", "Tek sayfa — projeksiyon cihazına"],
    ]
    write_rows(ws, r, ["Sayfa", "Açıklama"], guide, {1: 18, 2: 50})


# ═══════════════════════════════════════════════════════════════
# SAYFA A — GİRDİLER (model parametreleri)
# ═══════════════════════════════════════════════════════════════
def sheet_inputs(wb: Workbook) -> dict[str, str]:
    """Girdi hücreleri — dönen dict formül referansları için."""
    ws = wb.create_sheet("A-Girdiler")
    title_block(ws, 1, "MODEL GİRDİLERİ", "Sarı hücreleri değiştirin — Finans ve Duyarlılık sayfaları güncellenir")
    col_w(ws, {1: 32, 2: 18, 3: 40})

    inputs = [
        ("KUR — USD/TRY", 38, "API faturaları için"),
        ("ORTALAMA PAKET (₺/ay)", 7990, "Profesyonel paket ARPU"),
        ("KURULUM ÜCRETİ (₺)", 14900, "Tek seferlik onboarding"),
        ("AYLIK CHURN (%)", 4, "Örn. 4 → formülde %4 kullanılır"),
        ("SABİT GİDER (₺/ay)", 8_500, "Altyapı + hukuk + araçlar (dev hariç)"),
        ("DEĞİŞKEN GİDER (₺/klinik/ay)", 1350, "OpenAI + WhatsApp payı"),
        ("FREELANCE DEV (₺/ay)", 45000, "Yarı zamanlı geliştirici"),
        ("SATIŞ (₺/ay) — başlangıç ayı", 7, "Hangi aydan itibaren"),
        ("SATIŞ MAAŞ (₺/ay)", 30000, "Yarı zamanlı"),
        ("DESTEK (₺/ay) — başlangıç ayı", 10, "5+ klinik sonrası"),
        ("DESTEK MAAŞ (₺/ay)", 22000, "Part-time"),
        ("BAŞLANGIÇ NAKİT (₺)", 500000, "Yatırım / öz kaynak"),
        ("CAC (₺/yeni klinik)", 12000, "Satış + pazarlama"),
    ]

    header_row(ws, 4, ["Parametre", "Değer", "Açıklama"])
    refs = {}
    for i, (name, val, note) in enumerate(inputs, 5):
        ws.cell(i, 1, name).border = _border()
        c = ws.cell(i, 2, val)
        c.border = _border()
        c.fill = _fill(C["amber_l"])
        c.font = _f(bold=True)
        c.alignment = Alignment(horizontal="center")
        ws.cell(i, 3, note).border = _border()
        key = name.split("(")[0].strip().replace(" ", "_").upper()[:20]
        refs[key] = f"'A-Girdiler'!$B${i}"

    # Yeni klinik rampası (ay 1-24)
    r = 20
    ws.merge_cells(f"A{r}:C{r}")
    ws[f"A{r}"] = "YENİ KLİNİK RAMPASI (aylık eklenen ödeyen klinik)"
    ws[f"A{r}"].font = _f(bold=True, color=C["navy"])
    r += 1
    header_row(ws, r, ["Ay", "Yeni klinik", "Not"])
    ramp = [0, 0, 0, 0, 1, 0, 1, 1, 1, 2, 1, 1, 1, 1, 2, 1, 1, 1, 1, 1, 1, 1, 1, 1]
    notes = ["Pilot"] * 4 + ["İlk satış"] + [""] * 19
    for m in range(24):
        ri = r + 1 + m
        ws.cell(ri, 1, f"Ay {m+1}").border = _border()
        c = ws.cell(ri, 2, ramp[m])
        c.border = _border()
        c.fill = _fill(C["amber_l"])
        ws.cell(ri, 3, notes[m] if m < len(notes) else "").border = _border()

    refs["RAMP"] = f"'A-Girdiler'!$B$22:$B$45"
    return refs


# ═══════════════════════════════════════════════════════════════
# SAYFA — FİNANS (24 ay formüllü)
# ═══════════════════════════════════════════════════════════════
def sheet_finance(wb: Workbook):
    ws = wb.create_sheet("Finans")
    title_block(ws, 1, "24 AYLIK FİNANSAL MODEL", "Tüm tutarlar ₺ — formüller A-Girdiler sayfasına bağlı")
    col_w(ws, {1: 6, 2: 12, 3: 12, 4: 12, 5: 14, 6: 14, 7: 14, 8: 14, 9: 14, 10: 14})

    headers = ["Ay", "Yeni", "Churn", "Aktif", "MRR", "Kurulum", "Gider", "Net", "Kümülatif", "Burn/mo"]
    header_row(ws, 4, headers)

    for m in range(24):
        r = 5 + m
        ws.cell(r, 1, m + 1).border = _border()
        # Yeni klinik
        ws.cell(r, 2, f"='A-Girdiler'!B{22 + m}").border = _border()
        # Churn (önceki ay aktif * churn rate) — ay 1 churn 0
        if m == 0:
            ws.cell(r, 3, 0).border = _border()
        else:
            ws.cell(r, 3, f"=ROUND(D{r-1}*'A-Girdiler'!$B$8/100,2)").border = _border()
        # Aktif
        if m == 0:
            ws.cell(r, 4, f"=B{r}").border = _border()
        else:
            ws.cell(r, 4, f"=MAX(0,D{r-1}-C{r}+B{r})").border = _border()
        # MRR
        ws.cell(r, 5, f"=D{r}*'A-Girdiler'!$B$6").border = _border()
        ws.cell(r, 5).number_format = '#,##0'
        # Kurulum geliri
        ws.cell(r, 6, f"=B{r}*'A-Girdiler'!$B$7").border = _border()
        ws.cell(r, 6).number_format = '#,##0'
        # Gider
        ws.cell(r, 7, (
            f"='A-Girdiler'!$B$9+D{r}*'A-Girdiler'!$B$10+'A-Girdiler'!$B$11"
            f"+IF(A{r}>='A-Girdiler'!$B$12,'A-Girdiler'!$B$13,0)"
            f"+IF(A{r}>='A-Girdiler'!$B$14,'A-Girdiler'!$B$15,0)"
            f"+B{r}*'A-Girdiler'!$B$17"
        )).border = _border()
        ws.cell(r, 7).number_format = '#,##0'
        # Net
        ws.cell(r, 8, f"=E{r}+F{r}-G{r}").border = _border()
        ws.cell(r, 8).number_format = '#,##0'
        # Kümülatif
        if m == 0:
            ws.cell(r, 9, f"='A-Girdiler'!$B$16+H{r}").border = _border()
        else:
            ws.cell(r, 9, f"=I{r-1}+H{r}").border = _border()
        ws.cell(r, 9).number_format = '#,##0'
        # Burn (negatif net)
        ws.cell(r, 10, f"=IF(H{r}<0,-H{r},0)").border = _border()
        ws.cell(r, 10).number_format = '#,##0'
        if m % 2 == 1:
            for c in range(1, 11):
                ws.cell(r, c).fill = _fill(C["gray"])

    # Özet hücreleri (Dashboard referansları) — satır 33+
    s = 33
    labels = [
        ("Aktif klinik (ay 12)", "=D28"),
        ("Aktif klinik (ay 24)", "=D40"),
        ("MRR ay 12", "=E28"),
        ("MRR ay 24", "=E40"),
        ("Başabaş ayı", "=IFERROR(MATCH(TRUE,E5:E28>=G5:G28,0),\">12\")"),
        ("Toplam kurulum geliri Y1", "=SUM(F5:F16)"),
        ("Yıl 1 net (kümülatif ay 12 - başlangıç)", "=I16-'A-Girdiler'!$B$16"),
        ("Yıl 2 MRR (ay 24)", "=E40"),
        ("Ortalama aylık burn Y1", "=AVERAGE(J5:J16)"),
        ("Runway (ay) — başlangıç/burn", "=IF(J16>0,'A-Girdiler'!$B$17/J16,\">24\")"),
    ]
    ws.merge_cells(f"A{s}:B{s}")
    ws[f"A{s}"] = "MODEL ÖZETİ"
    ws[f"A{s}"].font = _f(bold=True, size=14, color=C["navy"])
    for i, (lab, formula) in enumerate(labels, s + 1):
        ws.cell(i, 1, lab).font = _f(bold=True)
        ws.cell(i, 1).border = _border()
        c = ws.cell(i, 2, formula)
        c.border = _border()
        c.fill = _fill(C["green_l"])
        c.number_format = '#,##0'
        # Dashboard refs use column N (14) for some - map explicitly
        if lab.startswith("Aktif klinik (ay 12)"):
            ws.cell(i, 14, formula)  # N33
        if lab.startswith("MRR ay 12"):
            ws.cell(i, 14, formula)  # will overwrite - use separate cells

    # Dashboard named outputs in column N
    ws["N33"] = "=D28"
    ws["N35"] = "=E28"
    ws["N37"] = '=IFERROR(MATCH(TRUE,E5:E28>=G5:G28,0)&". ay","12+")'
    ws["N39"] = "=I16-'A-Girdiler'!$B$16"
    ws["N42"] = "=AVERAGE(J5:J16)"
    ws["N43"] = '=IF(N42>0,\'A-Girdiler\'!$B$16/N42,">24")'
    for ref in ("N33", "N35", "N39", "N42"):
        ws[ref].number_format = '#,##0'

    # Koşullu biçim — net negatif kırmızı
    ws.conditional_formatting.add(
        "H5:H28",
        CellIsRule(operator="lessThan", formula=["0"], fill=_fill(C["red_l"]), font=_f(color=C["red"])),
    )
    ws.conditional_formatting.add(
        "H5:H28",
        CellIsRule(operator="greaterThan", formula=["0"], fill=_fill(C["green_l"])),
    )

    # Grafik 1 — MRR vs Gider
    chart1 = LineChart()
    chart1.title = "MRR vs Gider (24 Ay)"
    chart1.style = 10
    chart1.y_axis.title = "₺"
    chart1.x_axis.title = "Ay"
    chart1.width = 18
    chart1.height = 10
    data1 = Reference(ws, min_col=5, min_row=4, max_row=28)
    data2 = Reference(ws, min_col=7, min_row=4, max_row=28)
    cats = Reference(ws, min_col=1, min_row=5, max_row=28)
    chart1.add_data(data1, titles_from_data=True)
    chart1.add_data(data2, titles_from_data=True)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "A42")

    # Grafik 2 — Kümülatif nakit
    chart2 = LineChart()
    chart2.title = "Kümülatif Nakit"
    chart2.style = 11
    chart2.width = 18
    chart2.height = 10
    data3 = Reference(ws, min_col=9, min_row=4, max_row=28)
    chart2.add_data(data3, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "J42")


# ═══════════════════════════════════════════════════════════════
# DUYARLILIK ANALİZİ
# ═══════════════════════════════════════════════════════════════
def sheet_sensitivity(wb: Workbook):
    ws = wb.create_sheet("Duyarlılık")
    title_block(ws, 1, "DUYARLILIK ANALİZİ", "Paket fiyatı × klinik sayısı → yıllık gelir (₺)")
    col_w(ws, {1: 14})
    for i in range(2, 10):
        col_w(ws, {i: 12})

    prices = [4990, 5990, 6990, 7990, 8990, 9990, 11990, 14990]
    clinics = [3, 5, 8, 10, 15, 20, 25, 30]

    ws.cell(4, 1, "Klinik ↓ / Fiyat →").font = _f(bold=True)
    for j, p in enumerate(prices, 2):
        c = ws.cell(4, j, p)
        c.font = _f(bold=True, color=C["white"])
        c.fill = _fill(C["teal"])
        c.border = _border()
        c.number_format = '#,##0'

    for i, n in enumerate(clinics, 5):
        ws.cell(i, 1, n).font = _f(bold=True)
        ws.cell(i, 1).border = _border()
        ws.cell(i, 1).fill = _fill(C["teal_l"])
        for j, p in enumerate(prices, 2):
            annual = n * p * 12
            cell = ws.cell(i, j, annual)
            cell.number_format = '#,##0'
            cell.border = _border()

    ws.conditional_formatting.add(
        "B5:I12",
        ColorScaleRule(
            start_type="min", start_color="FEE2E2",
            mid_type="percentile", mid_value=50, mid_color="FEF3C7",
            end_type="max", end_color="DCFCE7",
        ),
    )

    r = 14
    title_block(ws, r, "Churn etkisi — 12. ay aktif klinik", "", "F")
    r += 2
    header_row(ws, r, ["Aylık churn", "8. ay aktif (başlangıç rampa)", "12. ay MRR (₺)"])
    churns = [0.02, 0.03, 0.04, 0.05, 0.06, 0.08]
    for i, ch in enumerate(churns, r + 1):
        ws.cell(i, 1, f"%{ch*100:.0f}").border = _border()
        # simplified illustrative
        active_8 = max(1, round(5 * (1 - ch) ** 4))
        ws.cell(i, 2, active_8).border = _border()
        ws.cell(i, 3, active_8 * 7990).border = _border()
        ws.cell(i, 3).number_format = '#,##0'

    r = 23
    write_rows(ws, r, ["Senaryo", "Varsayım", "12. ay MRR", "Yorum"],
               [
                   ["Kötü", "3 klinik, 5.990 ₺", 215_640, "Satış yavaş"],
                   ["Gerçekçi", "Finans modeli", "=Finans!E28", "Rampaya bağlı"],
                   ["İyi", "12 klinik, 7.990 ₺", 1_150_560, "Referans etkisi güçlü"],
                   ["Agresif", "20 klinik, 8.490 ₺", 2_037_600, "2 satışçı gerekir"],
               ], {1: 12, 2: 22, 3: 14, 4: 30})


# ═══════════════════════════════════════════════════════════════
# API MALİYET HESAPLAYICI
# ═══════════════════════════════════════════════════════════════
def sheet_api_calc(wb: Workbook):
    ws = wb.create_sheet("API Maliyet")
    title_block(ws, 1, "API MALİYET HESAPLAYICI", "Klinik büyüklüğüne göre aylık değişken maliyet")
    col_w(ws, {1: 28, 2: 14, 3: 14, 4: 14, 5: 30})

    tiers = [
        ("Küçük (1 hekim)", 200, 150, 50, 400),
        ("Orta (3 hekim) ★", 800, 500, 200, 1200),
        ("Büyük (6+ hekim)", 2000, 1200, 500, 2800),
        ("Zincir (3 şube)", 5000, 3000, 1200, 6500),
    ]
    write_rows(ws, 4, ["Profil", "WA konuşma/ay", "OpenAI token payı (₺)", "Diğer API (₺)", "Toplam (₺/ay)"],
               tiers, {1: 28, 2: 16, 3: 18, 4: 14, 5: 14})

    r = 11
    write_rows(ws, r, ["API", "Birim fiyat", "Tahmini kullanım", "Aylık (₺)", "Not"],
               [
                   ["OpenAI gpt-4o-mini", "~$0,15/1M token", "5M token/klinik", 285, f"USD/TRY={38}"],
                   ["WhatsApp Business (Meta)", "$0,05-0,15/konuşma", "500-2000", 1900, "Pazarlama vs hizmet"],
                   ["Vercel Pro (paylaşımlı)", "$20/ay", "Tüm platform", 760, "Sabit pay"],
                   ["VPS (paylaşımlı)", "170 ₺", "Tüm platform", 170, "10 klinikte ~17 ₺/klinik"],
                   ["İYS API", "Sabit + işlem", "Tüm platform", 500, "Ticari mesaj"],
                   ["E-posta bildirim", "Ücretsiz katman", "1000/ay", 0, "Resend"],
                   ["Google Business API", "Ücretsiz okuma", "—", 0, "Yayınlama ayrı"],
                   ["Supabase (self-host)", "VPS içi", "—", 0, "TR sunucu"],
               ], {1: 22, 2: 18, 3: 18, 4: 12, 5: 28})

    r = 22
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "Marj kontrolü: Profesyonel paket 7.990 ₺ — orta klinik API ~1.200 ₺ → brüt marj ~%85 (destek hariç)"
    ws[f"A{r}"].font = _f(bold=True, color=C["teal_d"])
    ws[f"A{r}"].fill = _fill(C["teal_l"])


# ═══════════════════════════════════════════════════════════════
# YOL HARİTASI + GANTT
# ═══════════════════════════════════════════════════════════════
def sheet_roadmap_gantt(wb: Workbook):
    ws = wb.create_sheet("Yol Haritası")
    title_block(ws, 1, "ÜRÜN & İŞ YOL HARİTASI", "2026 H2 — 2027 | Gantt: ■ = aktif çalışma")
    col_w(ws, {1: 4, 2: 22, 3: 28, 4: 22, 5: 10})
    for m in range(6, 18):
        col_w(ws, {m: 4})

    months = ["Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara", "Oca", "Şub", "Mar", "Nis", "May"]
    hdr = ["#", "Faz", "Klinik kazanır", "Biz yaparız", "Yatırım"] + months
    header_row(ws, 4, hdr)

    gantt_data = [
        [0, "F0 Güven", "Panel çalışır", "VPS, otomasyon, sağlık", "Düşük", 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        [1, "F1 Günlük iş", "WA yanıt, Google onay", "Inbox UI, bildirim", "Orta", 0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0],
        [2, "F2 Profesyonel", "Grafik, mobil, PDF", "UI redesign", "Orta", 0, 0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0],
        [3, "F3 Satış", "5 ödeyen klinik", "Çoklu klinik, İYS", "Yüksek", 0, 0, 0, 1, 1, 1, 1, 0, 0, 0, 0, 0],
        [4, "F4 Ölçek", "HBYS, resmi WA", "n8n→API geçiş", "Yüksek", 0, 0, 0, 0, 0, 0, 1, 1, 1, 1, 0, 0],
        [5, "F5 Platform", "10+ klinik", "SLA, yedek sunucu", "Orta", 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 1],
        [6, "KVKK/İYS", "Yasal güvence", "Hukuk, envanter", "Düşük", 0, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        [7, "Satış GTM", "Yeni müşteri", "Case study, demo", "Orta", 0, 0, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    ]

    for ri, row in enumerate(gantt_data, 5):
        for ci in range(1, 6):
            ws.cell(ri, ci, row[ci - 1]).border = _border()
        for mi, active in enumerate(row[5:], 6):
            cell = ws.cell(ri, mi)
            cell.border = _border()
            if active:
                cell.value = "■"
                cell.font = _f(bold=True, color=C["teal"])
                cell.fill = _fill(C["teal_l"])
                cell.alignment = Alignment(horizontal="center")

    r = 15
    write_rows(ws, r, ["Faz", "Bitti sayılır eğer…"],
               [
                   ["F0", "Medident yöneticisi haftada 5+ gün panele girer"],
                   ["F1", "Inbox yanıtlarının %80'i panelden gider"],
                   ["F2", "Mobil uyumlu; aylık PDF rapor tek tık"],
                   ["F3", "5 klinik aktif fatura kesiliyor"],
                   ["F4", "n8n olmadan inbox + NPS çalışıyor"],
                   ["F5", "Net gelir > operasyon giderinin 3 katı"],
               ], {1: 8, 2: 65})


# ═══════════════════════════════════════════════════════════════
# FİYATLANDIRMA + ÖZELLİK MATRİSİ
# ═══════════════════════════════════════════════════════════════
def sheet_pricing_matrix(wb: Workbook):
    ws = wb.create_sheet("Fiyatlandırma")
    title_block(ws, 1, "PAKETLER & ÖZELLİK MATRİSİ")

    features = [
        ("WhatsApp Inbox", 1, 1, 1, 1),
        ("Web chatbot", 1, 1, 1, 1),
        ("AI taslak yanıt", 1, 1, 1, 1),
        ("NPS / Feedback loop", 0, 1, 1, 1),
        ("Google yorum AI", 0, 1, 1, 1),
        ("İtibar (Sentinel)", 0, 1, 1, 1),
        ("Recall kampanya", 0, 1, 1, 1),
        ("eNPS (çalışan)", 0, 0, 1, 1),
        ("HBYS entegrasyon", 0, 0, 1, 1),
        ("Çoklu şube", 0, 0, 1, 1),
        ("SLA + öncelikli destek", 0, 0, 1, 1),
        ("White-label", 0, 0, 0, 1),
    ]
    header_row(ws, 4, ["Özellik", "Başlangıç\n4.990₺", "Profesyonel ★\n7.990₺", "Kurumsal\n14.990₺", "Enterprise"])
    for ri, (feat, *vals) in enumerate(features, 5):
        ws.cell(ri, 1, feat).border = _border()
        for ci, v in enumerate(vals, 2):
            cell = ws.cell(ri, ci, "✓" if v else "—")
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center")
            if v:
                cell.font = _f(bold=True, color=C["teal"])
                cell.fill = _fill(C["green_l"])

    r = 19
    write_rows(ws, r, ["Paket", "Aylık", "Kurulum", "Yıllık (indirimli)", "Hedef klinik"],
               [
                   ["Başlangıç", "4.990 ₺", "9.900 ₺", "50.900 ₺", "1 hekim"],
                   ["Profesyonel ★", "7.990 ₺", "14.900 ₺", "81.500 ₺", "Orta klinik"],
                   ["Kurumsal", "14.990 ₺", "29.900 ₺", "152.900 ₺", "Zincir"],
                   ["Enterprise", "Teklif", "Teklif", "Özel", "Hastane"],
               ], {1: 16, 2: 12, 3: 12, 4: 16, 5: 20})

    r = 25
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = (
        "ROI: Orta diş kliniği aylık 80-150 WhatsApp mesajı alır. Cevapsız %30 = 24-45 kayıp. "
        "1 implant randevusu (~20.000 ₺) Nefalix'i 2,5 yıl öder. Google yıldız +0,3 = %5-12 daha fazla arama görünürlüğü."
    )
    ws[f"A{r}"].alignment = Alignment(wrap_text=True)
    ws[f"A{r}"].fill = _fill(C["blue_l"])
    ws.row_dimensions[r].height = 48


# ═══════════════════════════════════════════════════════════════
# UYUMLULUK CHECKLIST
# ═══════════════════════════════════════════════════════════════
def sheet_compliance(wb: Workbook):
    ws = wb.create_sheet("Uyumluluk")
    title_block(ws, 1, "KVKK / İYS / SAĞLIK UYUMLULUK CHECKLIST", "Satış öncesi tamamlanması gerekenler")
    col_w(ws, {1: 6, 2: 32, 3: 12, 4: 12, 5: 36})

    items = [
        (1, "VERBİS kaydı veya güncelleme", "P0", "Yapılacak", "Veri sorumlusu olarak klinik + Nefalix iş ortaklığı netleştir"),
        (2, "Aydınlatma metni (hasta)", "P0", "Yapılacak", "Web + WhatsApp + form"),
        (3, "Açık rıza metni (ticari ileti)", "P0", "Yapılacak", "İYS öncesi şart"),
        (4, "Veri işleme envanteri", "P0", "Yapılacak", "Supabase tabloları listesi"),
        (5, "İYS kayıt (şirket)", "P1", "Yapılacak", "Ticari mesaj için"),
        (6, "Veri işleyen sözleşmesi (DPA)", "P1", "Yapılacak", "Klinik ↔ Nefalix"),
        (7, "Sunucu lokasyon belgesi (TR)", "P0", "Tamam", "Sunucumburada İstanbul"),
        (8, "OpenAI veri işleme (DPA)", "P1", "Yapılacak", "API — kişisel veri minimizasyonu"),
        (9, "Saklama süresi politikası", "P1", "Yapılacak", "Örn. 2 yıl sonra anonimleştir"),
        (10, "Veri ihlali müdahale planı", "P2", "Yapılacak", "72 saat KVKK bildirimi"),
        (11, "AI kullanım sınırı (tıbbi değil)", "P0", "Tamam", "Operasyonel/randevu odaklı"),
        (12, "WhatsApp ticari mesaj İYS kontrolü", "P1", "Yapılacak", "Workflow + API"),
    ]
    write_rows(ws, 4, ["#", "Madde", "Öncelik", "Durum", "Not"], items, {1: 5, 2: 32, 3: 10, 4: 12, 5: 38})

    for row in range(5, 5 + len(items)):
        st = ws.cell(row, 4).value
        if st == "Tamam":
            ws.cell(row, 4).fill = _fill(C["green_l"])
        else:
            ws.cell(row, 4).fill = _fill(C["amber_l"])
        pr = ws.cell(row, 3).value
        if pr == "P0":
            ws.cell(row, 3).fill = _fill(C["red_l"])


# ═══════════════════════════════════════════════════════════════
# UZMAN GÖRÜŞMELERİ (Q&A)
# ═══════════════════════════════════════════════════════════════
def sheet_experts(wb: Workbook):
    ws = wb.create_sheet("Uzman Görüşmeleri")
    title_block(ws, 1, "UZMAN GÖRÜŞMELERİ — SORU & CEVAP", "5 disiplin; yol haritası bu görüşmelerle revize edildi")
    col_w(ws, {1: 16, 2: 40, 3: 40})

    qa = [
        ("SaaS Mimarı", "n8n'i ne zaman bırakmalıyız?", "5 ödeyen klinikten sonra modül modül. Önce okuma/yazma API, sonra AI. n8n prototip olarak kalabilir."),
        ("SaaS Mimarı", "Tek VPS yeterli mi?", "8 kliniğe kadar evet. 10+ için DB ayrı sunucu veya yedekleme zorunlu."),
        ("KVKK Avukatı", "Hasta verisi yurt dışına çıkar mı?", "OpenAI'ye giden prompt'ta isim/telefon maskeleme. Vercel sadece oturum — hasta tablosu yok."),
        ("KVKK Avukatı", "Evolution WhatsApp yasal mı?", "Pilot tolere edilebilir; ticari ölçekte Meta resmi API + İYS şart."),
        ("Diş Kliniği Sahibi", "Buna neden para vereyim?", "Kaçan WhatsApp = kayıp implant. Google yorum = yeni hasta. Sekreter yükü azalır."),
        ("Diş Kliniği Sahibi", "Kurulum ne kadar sürer?", "1-2 gün uzaktan. QR + panel eğitimi 1 saat."),
        ("Satış Direktörü", "İlk 10 müşteriyi nasıl buluruz?", "Referans (Medident) + İstanbul yüz yüze + HBYS partner."),
        ("Satış Direktörü", "Fiyat direnci nerede?", "5.000 ₺ altı 'ucuz ve eksik' algısı. 8.000 ₺ orta klinik için sweet spot."),
        ("Finans CFO", "Yıl 1 neden zarar?", "SaaS normali. CAC + geliştirici sabit. Yıl 2'de marj açılır."),
        ("Finans CFO", "Kurulum ücreti şart mı?", "Evet — nakit akışı ve ciddiyet filtresi. 14.900 ₺ önerilir."),
        ("Ürün UX", "İlk ne geliştirilmeli?", "Inbox thread + Google onay. Grafikler sonra."),
        ("Ürün UX", "Mobil şart mı?", "Sekreter telefondan bakar — F2'de zorunlu."),
        ("DevOps", "İzleme ne zaman?", "5. klinikten sonra uptime alarmı. Pilot'ta manuel yeter."),
        ("DevOps", "Yedekleme?", "Günlük pg_dump + haftalık offsite. 10+ klinikte S3 TR."),
        ("Rakip Analisti", "Podium neden tehdit değil?", "TR yok, TL fiyat yok, KVKK uyumu yok. Yerel avantaj."),
    ]
    write_rows(ws, 4, ["Uzman", "Soru", "Cevap / Öneri"], qa, {1: 16, 2: 38, 3: 42})


# ═══════════════════════════════════════════════════════════════
# REKABET SKOR KARTI
# ═══════════════════════════════════════════════════════════════
def sheet_competition_score(wb: Workbook):
    ws = wb.create_sheet("Rekabet")
    title_block(ws, 1, "REKABET SKOR KARTI", "1=düşük 5=yüksek — Nefalix hedef pazar: orta diş kliniği TR")
    criteria = ["Fiyat", "AI", "WhatsApp", "Google yorum", "KVKK/TR", "Kolay kurulum", "Destek TR"]
    competitors = {
        "Nefalix": [4, 5, 5, 5, 5, 4, 5],
        "Yerel WA CRM": [5, 2, 4, 1, 3, 5, 4],
        "Bulut Klinik": [3, 1, 2, 2, 4, 3, 4],
        "Podium": [1, 4, 4, 5, 1, 2, 1],
        "Manuel": [5, 1, 3, 2, 5, 5, 5],
        "Ajans": [2, 2, 3, 4, 3, 2, 4],
    }
    header_row(ws, 4, ["Kriter"] + list(competitors.keys()))
    col_w(ws, {1: 16})
    for i in range(2, 2 + len(competitors)):
        col_w(ws, {i: 12})

    for ri, crit in enumerate(criteria, 5):
        ws.cell(ri, 1, crit).border = _border()
        ws.cell(ri, 1).font = _f(bold=True)
        for ci, (name, scores) in enumerate(competitors.items(), 2):
            v = scores[ri - 5]
            cell = ws.cell(ri, ci, v)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center")
            if name == "Nefalix":
                cell.fill = _fill(C["teal_l"])
                cell.font = _f(bold=True, color=C["teal_d"])

    ws.conditional_formatting.add(
        f"B5:{get_column_letter(1+len(competitors))}{4+len(criteria)}",
        ColorScaleRule(start_type="min", start_color="FEE2E2", end_type="max", end_color="DCFCE7"),
    )

    r = 14
    write_rows(ws, r, ["Segment", "Büyüklük", "Nefalix hedef"],
               [
                   ["TAM", "~25.000 diş kliniği TR", "Uzun vadeli"],
                   ["SAM", "~4.000 dijital olgun klinik", "3-5 yıl"],
                   ["SOM (3 yıl)", "150 klinik", "%3,75 SAM"],
                   ["SOM (1 yıl)", "8-10 klinik", "Gerçekçi pilot sonrası"],
               ], {1: 10, 2: 28, 3: 30})


# ═══════════════════════════════════════════════════════════════
# MÜŞTERİ YOLCULUĞU
# ═══════════════════════════════════════════════════════════════
def sheet_customer_journey(wb: Workbook):
    ws = wb.create_sheet("Müşteri Yolculuğu")
    title_block(ws, 1, "KLİNİK MÜŞTERİ YOLCULUĞU", "İlk temas → sadık müşteri")
    write_rows(ws, 4, ["Aşama", "Süre", "Klinik ne yaşar?", "Biz ne yaparız?", "Çıktı", "Maliyet"],
               [
                   ["1. Demo", "30 dk", "Canlı panel görür", "Medident verisiyle demo", "İlgi", "0"],
                   ["2. Teklif", "3-5 gün", "Paket + fiyat", "ROI hesabı sunarız", "Onay", "CAC başlar"],
                   ["3. Sözleşme", "1 hafta", "KVKK + DPA imza", "Hukuk metinleri", "Yasal OK", "~2.500 ₺"],
                   ["4. Kurulum", "1-2 gün", "QR, panel, eğitim", "Onboarding script", "Canlı", "8-12 saat iş"],
                   ["5. Go-live", "Hafta 1", "İlk mesajlar gelir", "Destek hattı", "İlk değer", "Destek"],
                   ["6. Başarı", "Ay 1-3", "NPS + yorum artışı", "Aylık check-in", "Case study", "0"],
                   ["7. Genişleme", "Ay 6+", "2. şube / referans", "Upsell kurumsal", "LTV artışı", "Düşük CAC"],
               ], {1: 12, 2: 10, 3: 24, 4: 24, 5: 16, 6: 12})


# ═══════════════════════════════════════════════════════════════
# EKİP PLANI
# ═══════════════════════════════════════════════════════════════
def sheet_team(wb: Workbook):
    ws = wb.create_sheet("Ekip Planı")
    title_block(ws, 1, "EKİP & İŞE ALIM PLANI")
    write_rows(ws, 4, ["Rol", "Ne zaman?", "Zaman", "Aylık (₺)", "Görev"],
               [
                   ["Kurucu / CEO", "Şimdi", "Full", "Hisse", "Vizyon, satış, ürün"],
                   ["Freelance developer", "Tem 2026", "%50", "45.000", "Panel + API geçişi"],
                   ["Hukuk (KVKK)", "Tem 2026", "Proje", "25.000 (tek)", "Sözleşmeler"],
                   ["Satış / BD", "Eyl 2026", "%50", "30.000", "Demo, kapanış"],
                   ["Müşteri desteği", "10. klinik", "%50", "22.000", "Ticket, eğitim"],
                   ["Full-time developer", "20. klinik", "Full", "80.000", "Ölçek, HBYS"],
               ], {1: 20, 2: 12, 3: 10, 4: 14, 5: 36})


# ═══════════════════════════════════════════════════════════════
# MEDIDENT PİLOT SKOR KARTI
# ═══════════════════════════════════════════════════════════════
def sheet_pilot_scorecard(wb: Workbook):
    ws = wb.create_sheet("Medident Pilot")
    title_block(ws, 1, "MEDIDENT İSTANBUL — PİLOT SKOR KARTI", "clinic_id: 51738ea8-c12e-40ce-a0e2-42869496d76b")
    write_rows(ws, 4, ["Metrik", "Hedef", "Şu an", "Durum"],
               [
                   ["Panel erişilebilirlik", "%99", "VPS askı", "🔴"],
                   ["Haftalık panel girişi (yönetici)", "5+/hafta", "—", "⏳"],
                   ["Inbox yanıt panelden", "%80", "—", "⏳"],
                   ["Google yorum onay oranı", "%50", "—", "⏳"],
                   ["WhatsApp bağlı", "Evet", "QR gerekli", "🟡"],
                   ["NPS ortalama", ">8", "—", "⏳"],
                   ["Google puan değişimi", "+0,2", "—", "⏳"],
                   ["Referans hazır (case study)", "Evet", "Hayır", "🟡"],
               ], {1: 28, 2: 14, 3: 14, 4: 10})


# ═══════════════════════════════════════════════════════════════
# YATIRIMCI METRİKLERİ
# ═══════════════════════════════════════════════════════════════
def sheet_investor(wb: Workbook):
    ws = wb.create_sheet("Yatırımcı")
    title_block(ws, 1, "YATIRIMCI / ORTAK ÖZETİ", "SaaS metrikleri — formüller Finans sayfasına bağlı")
    write_rows(ws, 4, ["Metrik", "Değer", "Formül / kaynak", "Sektör benchmark"],
               [
                   ["MRR (ay 12)", "=Finans!E28", "Aktif × ARPU", "Erken stage: 50K+ ₺ iyi"],
                   ["ARR (ay 12)", "=Finans!E28*12", "MRR × 12", "—"],
                   ["Brüt marj", "%82", "Birim ekonomi", "SaaS: >70%"],
                   ["CAC", "12.000 ₺", "Girdiler", "B2B SMB: 5-20K ₺"],
                   ["LTV (18 ay)", "143.820 ₺", "7990×18", "—"],
                   ["LTV/CAC", "12x", "LTV/CAC", ">3x sağlıklı"],
                   ["Churn", "%4/ay", "Girdiler", "SMB: 3-7%"],
                   ["Payback", "~1,5 ay", "CAC/ARPU", "<12 ay iyi"],
                   ["Burn (ort Y1)", "=Finans!N42", "Finans modeli", "—"],
                   ["Runway", "=Finans!N43", "Nakit/burn", ">12 ay güvenli"],
               ], {1: 18, 2: 16, 3: 28, 4: 28})

    r = 16
    write_rows(ws, r, ["Kilometre taşı", "Tarih", "Değerleme etkisi"],
               [
                   ["Medident pilot canlı", "Tem 2026", "Teknik risk düşer"],
                   ["3 ödeyen klinik", "Ara 2026", "PMF sinyali"],
                   ["8 ödeyen klinik", "Haz 2027", "Seed hazırlığı"],
                   ["n8n → API tamam", "Eyl 2027", "Ölçeklenebilirlik"],
                   ["50 klinik", "2028", "Series A adayı"],
               ], {1: 24, 2: 12, 3: 36})


# ═══════════════════════════════════════════════════════════════
# SUNUM SAYFASI (tek sayfa projeksiyon)
# ═══════════════════════════════════════════════════════════════
def sheet_presentation(wb: Workbook):
    ws = wb.create_sheet("Sunum")
    ws.sheet_view.showGridLines = False
    col_w(ws, {1: 2, 2: 20, 3: 20, 4: 20, 5: 20, 6: 2})

    for row in range(1, 45):
        for col in range(1, 7):
            ws.cell(row, col).fill = _fill(C["navy"])

    ws.merge_cells("B2:E2")
    ws["B2"] = "NEFALIX AI"
    ws["B2"].font = _f(bold=True, size=28, color=C["white"])
    ws.merge_cells("B3:E3")
    ws["B3"] = "Diş Klinikleri İçin Hasta İletişim & İtibar Platformu"
    ws["B3"].font = _f(size=14, color=C["teal_l"])

    boxes = [
        ("B5", "SORUN", "Klinikler WhatsApp mesajını kaçırıyor,\nyorumlara geç yanıt veriyor."),
        ("C5", "ÇÖZÜM", "Tek panel: inbox, NPS, Google AI,\nitibar takibi — Türkiye sunucu."),
        ("D5", "PAZAR", "4.000 dijital olgun klinik (TR)\nHedef: 150 klinik / 3 yıl"),
        ("E5", "TRAKSİYON", "Medident İstanbul pilot\nPanel + 8 modül kurulu"),
    ]
    for cell, title, body in boxes:
        c0, r0 = cell[0], int(cell[1:])
        ws.merge_cells(f"{c0}{r0}:{c0}{r0+3}")
        ws[cell] = f"{title}\n\n{body}"
        ws[cell].font = _f(size=10, color=C["white"])
        ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
        ws[cell].fill = _fill(C["teal_d"])

    ws.merge_cells("B10:E10")
    ws["B10"] = "İŞ MODELİ"
    ws["B10"].font = _f(bold=True, size=14, color=C["gold"])

    pricing = "Başlangıç 4.990 ₺  |  Profesyonel 7.990 ₺  |  Kurumsal 14.990 ₺  |  + kurulum ücreti"
    ws.merge_cells("B11:E11")
    ws["B11"] = pricing
    ws["B11"].font = _f(size=11, color=C["white"])

    ws.merge_cells("B13:E13")
    ws["B13"] = "12 AY HEDEF"
    ws["B13"].font = _f(bold=True, size=14, color=C["gold"])
    ws.merge_cells("B14:E14")
    ws["B14"] = "=Finans!N33&\" klinik  |  \"&TEXT(Finans!N35,\"#,##0\")&\" ₺ MRR  |  Yıl 1 yatırım: ~2,1M ₺\""
    ws["B14"].font = _f(size=12, color=C["white"])

    ws.merge_cells("B16:E16")
    ws["B16"] = "YATIRIM İHTİYACI"
    ws["B16"].font = _f(bold=True, size=14, color=C["gold"])
    ws.merge_cells("B17:E18")
    ws["B17"] = (
        "500K ₺ başlangıç nakit (mevcut) + opsiyonel 1-1,5M ₺ seed\n"
        "Kullanım: geliştirici, satış, hukuk, altyapı, 18 ay runway"
    )
    ws["B17"].font = _f(size=11, color=C["white"])
    ws["B17"].alignment = Alignment(wrap_text=True)

    ws.merge_cells("B20:E20")
    ws["B20"] = "SORULACAK ONAYLAR"
    ws["B20"].font = _f(bold=True, size=14, color=C["gold"])
    approvals = "☐ Yıllık bütçe onayı  ☐ Medident referans izni  ☐ KVKK hukuk bütçesi  ☐ Freelance dev işe alım"
    ws.merge_cells("B21:E21")
    ws["B21"] = approvals
    ws["B21"].font = _f(size=10, color=C["teal_l"])

    ws.merge_cells("B23:E23")
    ws["B23"] = f"nefalixai.com | {TODAY.strftime('%d.%m.%Y')} | Gizli"
    ws["B23"].font = _f(size=9, color=C["slate"], italic=True)


# ═══════════════════════════════════════════════════════════════
# AKSİYONLAR + VARSAYIMLAR + ESKİ SAYFALAR (kısa)
# ═══════════════════════════════════════════════════════════════
def sheet_actions(wb: Workbook):
    ws = wb.create_sheet("Aksiyonlar")
    title_block(ws, 1, "180 GÜNLÜK AKSİYON PLANI")
    rows = [
        [1, "VPS askı kaldır", "Teknik", "Haz 26", 0, "P0", "Blokaj"],
        [2, "KVKK paketi", "Hukuk", "Tem 26", 25000, "P0", "Satış öncesi"],
        [3, "Medident eğitim", "Ops", "Tem 26", 0, "P0", "Adoption"],
        [4, "Sağlık göstergesi UI", "Dev", "Tem 26", 0, "P1", "Güven"],
        [5, "Google onay ekranı", "Dev", "Ağu 26", 0, "P1", "Değer"],
        [6, "Inbox thread UI", "Dev", "Ağu 26", 0, "P1", "Değer"],
        [7, "İYS kayıt", "Ops", "Ağu 26", 6000, "P1", "Yasal"],
        [8, "Case study video", "Satış", "Eyl 26", 5000, "P1", "GTM"],
        [9, "2. pilot klinik", "Satış", "Eki 26", 0, "P2", "PMF"],
        [10, "Mobil responsive", "Dev", "Eki 26", 0, "P2", "UX"],
        [11, "WA Business API", "Teknik", "Kas 26", 5000, "P2", "Ölçek"],
        [12, "API v1 (Supabase read)", "Dev", "Ara 26", 0, "P2", "Mimari"],
        [13, "3. ödeyen klinik", "Satış", "Ara 26", 0, "P1", "Gelir"],
        [14, "Yedekleme otomasyon", "DevOps", "Oca 27", 400, "P2", "Risk"],
        [15, "Yıllık bütçe revizyon", "Yönetim", "Haz 27", 0, "P0", "Plan"],
    ]
    end = write_rows(ws, 4, ["#", "Aksiyon", "Sorumlu", "Tarih", "₺", "P", "Etki"], rows,
                     {1: 4, 2: 32, 3: 10, 4: 10, 5: 10, 6: 6, 7: 14})
    for row in range(5, 5 + len(rows)):
        p = ws.cell(row, 6).value
        if p == "P0":
            ws.cell(row, 6).fill = _fill(C["red_l"])


def sheet_assumptions(wb: Workbook):
    ws = wb.create_sheet("Varsayımlar")
    title_block(ws, 1, "VARSAYIMLAR & SORUMLULUK REDDİ")
    items = [
        "Finans modeli A-Girdiler sayfasındaki sarı hücrelere bağlıdır — değiştirince tüm tablo güncellenir.",
        f"USD/TRY = 38 (Haziran 2026 tahmini). Kur artışı API maliyetini artırır.",
        "Churn %4/ay ilk yıl — iyileştirme ile %2,5'a düşürülebilir.",
        "Medident pilot ücretsiz — gelir modeline dahil değil.",
        "OpenAI ve Meta fiyatları değişkendir — API Maliyet sayfasına bakın.",
        "Yıl 1 zarar SaaS'ta normaldir; yatırım yılı olarak planlayın.",
        "Hukuki metinler avukat onayı gerektirir — bu rapor hukuki tavsiye değildir.",
        "Rakip skorları çerçeve analizidir — birincil kaynak değildir.",
        "Uzman Q&A sentetik çerçeve analizidir — resmi danışmanlık yerine geçmez.",
        "VERBİS/İYS süreçleri klinik + Nefalix ortak sorumluluğundadır.",
    ]
    for i, t in enumerate(items, 4):
        ws.cell(i, 1, i - 3).border = _border()
        ws.merge_cells(f"B{i}:F{i}")
        ws.cell(i, 2, t).alignment = Alignment(wrap_text=True)
    col_w(ws, {1: 4, 2: 80})


def sheet_kpi_monthly(wb: Workbook):
    ws = wb.create_sheet("KPI Takip")
    title_block(ws, 1, "AYLIK KPI TAKİP ŞABLONU", "Her ay doldurun — yönetim kurulu toplantısı")
    months = ["Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara", "Oca", "Şub", "Mar", "Nis", "May", "Haz"]
    hdr = ["KPI"] + months + ["Hedef"]
    header_row(ws, 4, hdr)
    kpis = [
        ("Aktif klinik", [0]*12, 8),
        ("MRR (₺)", [0]*12, 64000),
        ("Yeni klinik", [0]*12, "—"),
        ("Churn (klinik)", [0]*12, 0),
        ("Panel kullanım %", [0]*12, 85),
        ("Inbox yanıt süresi (saat)", [0]*12, 2),
        ("NPS ortalama", [0]*12, 8.5),
        ("Destek ticket", [0]*12, 10),
        ("Uptime %", [0]*12, 99.5),
    ]
    for ri, (name, vals, target) in enumerate(kpis, 5):
        ws.cell(ri, 1, name).border = _border()
        ws.cell(ri, 1).font = _f(bold=True)
        for mi in range(12):
            c = ws.cell(ri, 2 + mi, vals[mi] if vals[mi] else "")
            c.border = _border()
            c.fill = _fill(C["amber_l"])
        ws.cell(ri, 14, target).border = _border()
        ws.cell(ri, 14).fill = _fill(C["green_l"])


def add_pie_chart_costs(wb: Workbook):
    """Maliyet dağılımı pasta grafiği — Finans sayfasına ek."""
    ws = wb.create_sheet("Maliyet Dağılımı")
    title_block(ws, 1, "AYLIK MALİYET DAĞILIMI (satışa hazır senaryo)")
    data = [
        ("VPS + altyapı", 620),
        ("Vercel", 760),
        ("OpenAI", 1500),
        ("WhatsApp Meta", 2000),
        ("Hukuk/KVKK", 2500),
        ("İYS", 500),
        ("Araçlar", 800),
        ("Freelance dev", 45000),
    ]
    write_rows(ws, 4, ["Kalem", "₺/ay"], data, {1: 24, 2: 12})

    pie = PieChart()
    pie.title = "Maliyet Dağılımı"
    pie.width = 14
    pie.height = 10
    labels = Reference(ws, min_col=1, min_row=5, max_row=4 + len(data))
    values = Reference(ws, min_col=2, min_row=5, max_row=4 + len(data))
    pie.add_data(values)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "D4")


def main():
    wb = Workbook()
    sheet_inputs(wb)
    sheet_finance(wb)
    sheet_dashboard(wb)  # active after finance for formulas
    sheet_sensitivity(wb)
    sheet_api_calc(wb)
    sheet_roadmap_gantt(wb)
    sheet_pricing_matrix(wb)
    sheet_compliance(wb)
    sheet_experts(wb)
    sheet_competition_score(wb)
    sheet_customer_journey(wb)
    sheet_team(wb)
    sheet_pilot_scorecard(wb)
    sheet_investor(wb)
    sheet_presentation(wb)
    sheet_actions(wb)
    sheet_kpi_monthly(wb)
    add_pie_chart_costs(wb)
    sheet_assumptions(wb)

    # Sayfa sırasını düzenle
    order = [
        "0-Dashboard", "Sunum", "A-Girdiler", "Finans", "Duyarlılık",
        "API Maliyet", "Maliyet Dağılımı", "Fiyatlandırma", "Yol Haritası",
        "Uyumluluk", "Uzman Görüşmeleri", "Rekabet", "Müşteri Yolculuğu",
        "Ekip Planı", "Medident Pilot", "Yatırımcı", "KPI Takip",
        "Aksiyonlar", "Varsayımlar",
    ]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"✓ Gelişmiş rapor ({len(wb.sheetnames)} sayfa): {OUT}")


if __name__ == "__main__":
    main()
