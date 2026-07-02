#!/usr/bin/env python3
"""Nefalix MVP Yol Haritası — renkli Excel üreteci"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
import os

OUTPUT = os.path.join(os.path.dirname(__file__), "../docs/Nefalix_MVP_Yol_Haritasi.xlsx")

# ── Renkler ──────────────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"   # başlık arkaplanı
GOLD        = "C9A84C"   # başlık yazı
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F5F5"

SPRINT_COLORS = {
    "Sprint 1 — Güvenlik Temizliği":     ("E8F4FD", "1A5276"),
    "Sprint 2 — Gerçek Auth Katmanı":    ("EAF6FF", "154360"),
    "Sprint 3 — Ödeme Sistemi":          ("FEF9E7", "7D6608"),
    "Sprint 4 — WhatsApp Üretim Geçişi": ("FDF2F8", "6C3483"),
    "Sprint 5 — Dashboard Modernizasyonu":("E9F7EF","145A32"),
    "Sprint 6 — n8n Stabilite ve İzleme":("FDF2E9","784212"),
    "Sprint 7 — Çoklu Entegrasyon":      ("FDFEFE","1B2631"),
    "Sprint 8 — Raporlama ve Analitik":  ("F4ECF7","512E5F"),
    "Gelecek Faz — Otel & Diğer Sektörler": ("EAFAF1","1E8449"),
    "Gelecek Faz — GEO / AEO (Pazarlama Vizyonu)": ("FDEDEC","922B21"),
}

ZORLUK_COLORS = {
    "Kolay": ("D5F5E3", "1E8449"),
    "Orta":  ("FEF9E7", "9A7D0A"),
    "Zor":   ("FADBD8", "922B21"),
}

ROWS = [
    # (sprint, iş, açıklama, araç, zorluk)
    ("Sprint 1 — Güvenlik Temizliği",
     "Supabase RLS politikalarını düzelt: anon key ile hasta, NPS, inbox tablolarına erişimi kapat",
     "Şu an herkes doğru URL'yi bilse hasta verilerini okuyabilir; bu KVKK ihlali",
     "Supabase SQL, supabase/migrations/, Row Level Security",
     "Orta"),
    ("Sprint 1 — Güvenlik Temizliği",
     "n8n webhook'larına kimlik doğrulama ekle (shared secret header)",
     "Dışarıdan biri webhook URL'ini bilirse NPS tetikleyebilir ya da sistemi karıştırabilir",
     "n8n HTTP Header Auth, .env secret",
     "Kolay"),
    ("Sprint 1 — Güvenlik Temizliği",
     "VPS güvenlik duvarı kurallarını kalıcı hale getir (reboot'ta sıfırlanmasın)",
     "Sunucu yeniden başladığında veritabanı portu internete açılıyor",
     "netfilter-persistent veya cron @reboot iptables-restore",
     "Kolay"),
    ("Sprint 1 — Güvenlik Temizliği",
     "Evolution API erişim bilgilerini döndür, eski credential'ları geçersiz kıl",
     "Eski şifre bir Cursor transcript'ine kaydolmuş; kötü niyetli biri sisteme girebilir",
     "VPS .env, n8n UI Settings, Evolution API",
     "Kolay"),

    ("Sprint 2 — Gerçek Auth Katmanı",
     "dashboard_users tablosu üzerine JWT tabanlı oturum sistemi kur",
     "Şu anki cookie auth geçici; birden fazla klinik geldiğinde herkes birbirinin verisini görebilir",
     "Supabase Auth veya custom JWT (jose/jsonwebtoken), Next.js API routes",
     "Orta"),
    ("Sprint 2 — Gerçek Auth Katmanı",
     "Rol yönetimi: klinik yöneticisi sadece kendi kliniğinin datasını görsün",
     "SaaS'ta her müşteri bağımsız bir 'kiracı'; veriler karışmamalı",
     "Supabase RLS clinic_id = auth.uid() policy",
     "Orta"),
    ("Sprint 2 — Gerçek Auth Katmanı",
     "Şifre sıfırlama, e-posta doğrulama akışı",
     "Müşteri kendi şifresini değiştiremezse destek talebine boğulunur",
     "Supabase Auth email templates veya Resend/SendGrid",
     "Kolay"),

    ("Sprint 3 — Ödeme Sistemi",
     "PayTR webhook'unu production'da test et ve idempotent hale getir",
     "Ödeme onayı iki kez gelirse abonelik iki kez açılmamalı; banka ödemeyi onaylamadan abonelik başlamamalı",
     "PayTR IPN, stripe_webhook_events tablosu (mevcut), HMAC doğrulama",
     "Zor"),
    ("Sprint 3 — Ödeme Sistemi",
     "Abonelik plan limitleri: Free/Growth/Pro katmanlarına göre klinik başına kota uygula",
     "Ücretsiz plan 100 NPS, ücretli plan sınırsız gibi kurallar olmadan SaaS çalışmaz",
     "n8n workflow içinde kota kontrolü + Supabase clinics.plan kolonu",
     "Orta"),
    ("Sprint 3 — Ödeme Sistemi",
     "Fatura sayfası: müşteri mevcut planını görsün, yükseltme yapabilsin",
     "Dashboard'daki billing sekmesi şu an statik; gerçek ödeme durumu yansıtılmıyor",
     "PayTR API + Supabase clinics tablosu",
     "Orta"),

    ("Sprint 4 — WhatsApp Üretim Geçişi",
     "Evolution API (pilot) yerine Meta resmi WABA API entegrasyonu",
     "Evolution ile 1000'den fazla mesaj atılırsa hesap yasaklanma riski var; resmi API şart",
     "Meta Business API, nefalix-00-whatsapp-gateway.json workflow'u revize",
     "Zor"),
    ("Sprint 4 — WhatsApp Üretim Geçişi",
     "İYS uyumluluğu: mesaj göndermeden önce izin kontrolü",
     "Türk mevzuatı: izinsiz ticari mesaj ceza gerektirir; iys_consent kolonu mevcut ama kontrol mekanizması yok",
     "Supabase patients.iys_consent + n8n'de filtre node'u",
     "Orta"),

    ("Sprint 5 — Dashboard Modernizasyonu",
     "HTML/CSS/JS dashboard'u React + Next.js'e taşı",
     "Mevcut dashboard düz HTML; yeni özellik eklemek giderek zorlaşıyor, test edilemiyor",
     "Next.js, Tailwind CSS, Supabase JS client",
     "Zor"),
    ("Sprint 5 — Dashboard Modernizasyonu",
     "Supabase Realtime ile canlı veri akışı",
     "Yeni bir NPS cevabı geldiğinde sayfayı yenilemeden ekranda görünsün",
     "Supabase Realtime subscription, React state",
     "Orta"),
    ("Sprint 5 — Dashboard Modernizasyonu",
     "Yeni klinik onboarding sihirbazı (wizard)",
     "Yazılımcıya ihtiyaç duymadan yeni müşteri eklenebilsin; şu an SQL migration ile ekleniyor",
     "Multi-step form, clinics INSERT + dashboard_users INSERT, setup token akışı",
     "Orta"),

    ("Sprint 6 — n8n Stabilite ve İzleme",
     "Başarısız workflow'lar için retry ve dead-letter mekanizması",
     "Supabase geçici çöktüğünde NPS cevabı kaybolmasın; yeniden denenmesi için kuyruk şart",
     "n8n Error Workflow + automation_events tablosu",
     "Orta"),
    ("Sprint 6 — n8n Stabilite ve İzleme",
     "Workflow execution izleme dashboard'u",
     "Kaç NPS gönderildi, kaçı başarısız — şu an smoke test elle çalıştırılıyor",
     "n8n Execution logs API + Supabase automation_events",
     "Kolay"),
    ("Sprint 6 — n8n Stabilite ve İzleme",
     "n8n'i worker modunda çalıştır (ölçeklenme)",
     "Aynı anda 10 klinikten NPS geldiğinde tek instance tıkayabilir",
     "n8n Queue Mode, Redis, VPS'de docker-compose güncellemesi",
     "Zor"),

    ("Sprint 7 — Çoklu Entegrasyon",
     "Medicasimple ve diğer klinik yazılımları için CRM adaptör şablonu",
     "Her yeni klinik entegrasyonu sıfırdan yazılmasın; Estesoft adaptörü şablon alınarak genişletilsin",
     "Mevcut nefalix-12-estesoft-adapter.json + nefalix-13-estesoft-poll.json şablon olarak",
     "Orta"),
    ("Sprint 7 — Çoklu Entegrasyon",
     "Google Yorum Yanıtlama otomasyonunu production'da aktif et",
     "AI tarafından hazırlanan yorum yanıtları şu an onay bekliyor ama tek tık onay arayüzü yok",
     "nefalix-11-google-review-approve.json + dashboard UI butonu",
     "Kolay"),

    ("Sprint 8 — Raporlama ve Analitik",
     "Klinik bazlı haftalık e-posta raporu",
     "Yönetici her Pazartesi 'geçen hafta 23 NPS, 4 olumsuz yorum' mailini otomatik alsın",
     "n8n Schedule Trigger + SendGrid/Resend + Supabase aggregation sorgusu",
     "Orta"),
    ("Sprint 8 — Raporlama ve Analitik",
     "Excel/PDF export (NPS, Google yorum skoru)",
     "Kliniklerin muhasebe veya raporlama toplantılarında veri çıktısı alabilmesi",
     "n8n CSV node + e-posta eki veya Vercel Edge Function",
     "Kolay"),

    ("Gelecek Faz — Otel & Diğer Sektörler",
     "Cloudbeds PMS entegrasyonu (otel sektörü)",
     "Checkout sonrası misafir otomatik NPS alsın — klinik yerine otel",
     "directives/integrations_hotels.md SOP mevcut, kod yazılmadı",
     "Orta"),
    ("Gelecek Faz — Otel & Diğer Sektörler",
     "Sector-agnostic çok kiracılı mimari (klinik/otel/oto)",
     "Aynı platform farklı sektöre hizmet verebilsin — clinics.sector kolonu mevcut",
     "Supabase sector enum + sector-aware workflow routing",
     "Zor"),

    ("Gelecek Faz — GEO / AEO (Pazarlama Vizyonu)",
     "AI Arama Motoru Optimizasyonu (GEO/AEO) katmanı",
     "Yapay Zeka araçları (ChatGPT, Perplexity) kliniği kaynak olarak göstersin — şimdilik gelecek vizyon, kurulmayacak",
     "Structured data, MCP Server, embedding tabanlı içerik",
     "Zor"),
    ("Gelecek Faz — GEO / AEO (Pazarlama Vizyonu)",
     "Nefalix MCP Server (klinik verisi için AI ağ geçidi)",
     "Yapay Zeka asistanları doğrudan klinik datasına bağlanabilsin — gelecek vizyon, MVP kapsamı dışı",
     "Model Context Protocol, klinik API, OAuth",
     "Zor"),
]

# ── Yardımcı fonksiyonlar ────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def bold_font(size=11, color="000000"):
    return Font(bold=True, size=size, color=color)

def normal_font(size=10, color="000000"):
    return Font(size=size, color=color)


# ── Workbook ─────────────────────────────────────────────────────────────────

wb = Workbook()

# ── SAYFA 1: Yönetici Özeti ──────────────────────────────────────────────────

ws1 = wb.active
ws1.title = "Yönetici Özeti"
ws1.sheet_view.showGridLines = False

# Başlık bandı
ws1.merge_cells("A1:D1")
ws1["A1"] = "Nefalix AI — MVP Yol Haritası"
ws1["A1"].font = Font(bold=True, size=20, color=GOLD)
ws1["A1"].fill = fill(DARK_NAVY)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 50

ws1.merge_cells("A2:D2")
ws1["A2"] = "Haziran 2026 | Pilot: MediDent Kartal | nefalixai.com"
ws1["A2"].font = Font(size=11, color="AAAAAA", italic=True)
ws1["A2"].fill = fill(DARK_NAVY)
ws1["A2"].alignment = Alignment(horizontal="center")
ws1.row_dimensions[2].height = 22

ws1.row_dimensions[3].height = 10

# Alt başlık
def section_header(ws, row, text):
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = text
    ws[f"A{row}"].font = Font(bold=True, size=13, color=WHITE)
    ws[f"A{row}"].fill = fill("1A3C5E")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28

def bullet_row(ws, row, text, bg="F8F9FA"):
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "  •  " + text
    ws[f"A{row}"].font = normal_font(11)
    ws[f"A{row}"].fill = fill(bg)
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    ws.row_dimensions[row].height = 40

section_header(ws1, 4, "✅  Neler Yapıldı ve Başarıldı?")
bullets_done = [
    "Hasta NPS akışı uçtan uca çalışıyor: Estesoft kliniği bağlandı → randevu bitişinde otomatik WhatsApp → puan 8-10: Google yorum linki → 7 altı: şikayet formu + yönetici WhatsApp alarmı.",
    "Google yorum senkronu: Google Places API'den yorumlar çekiliyor, Gemini AI ile yanıt taslağı oluşturuluyor, klinik yöneticisi onaylayıp yayınlıyor.",
    "İtibar izleme (Sentinel): Şikayetvar ve web'den klinik hakkında yazılanlar taranıyor, kritik olanlar anında WhatsApp'a iletiyor.",
    "Çalışan Memnuniyeti (eNPS) ve Kayıp Hasta Geri Çağırma modülleri kurulu.",
    "Gelen mesaj kutusu (Inbox) aktif; WhatsApp mesajları düşüyor, ekip içi yönlendirme çalışıyor.",
    "13 tablo, 20 Supabase migration — hastalar, randevular, yorumlar, NPS, şikayet izleme, faturalama tek veritabanında.",
    "Pilot klinik MediDent Kartal canlı veriyle test edildi.",
    "nefalixai.com canlıda — mobil uyumlu, video bölümü, fiyatlandırma, demo akışı var.",
    "Yönetici dashboard'u (KPI'lar, gelen kutusu, NPS kriz takibi, faturalama) Vercel'de yayında.",
    "PayTR ödeme entegrasyonu kod olarak hazır; canlı test bekleniyor.",
]
for i, b in enumerate(bullets_done):
    bg = "EBF5FB" if i % 2 == 0 else "F8F9FA"
    bullet_row(ws1, 5 + i, b, bg)

section_header(ws1, 16, "🧠  Neleri Öğrendik ve Tecrübe Ettik?")
bullets_learn = [
    "n8n'in drag-and-drop yapısı mantığı hızla kurmamızı sağladı ama smoke testlerde 7 iş akışının tamamı başarısız döndü; 'kuruldu' ile 'güvenilir çalışıyor' arasında ciddi fark var.",
    "Docker networking hatası (host.docker.internal vs localhost) birkaç oturumu bitirdi; artık direktif olarak yazılı.",
    "VPS'te güvenlik duvarı kuralları sunucu reboot'unda sıfırlanıyor — üretim için risk.",
    "Cursor ile fikrden çalışan koda 15-30 dakikada ulaşabildik. Ancak her oturumda bağlamı sıfırdan yükleyen AI için handoff + direktif sistemi kurmak zorunda kaldık.",
    "Smoke testler Cursor çıktısına körce güvenmek yerine bizim koyduğumuz en önemli kalite güvencesi oldu.",
]
for i, b in enumerate(bullets_learn):
    bg = "FEF9E7" if i % 2 == 0 else "FFFDE7"
    bullet_row(ws1, 17 + i, b, bg)

section_header(ws1, 23, "🚧  'Artık Yazılımcı Şart' Dediğimiz Kırılma Noktaları")
limits = [
    ("Güvenlik (KVKK)", "Hasta verileri anonim erişime açık", "Sistematik RLS revizyonu ve hukuki sorumluluk testi profesyonel göz ister"),
    ("Gerçek Kullanıcı Auth", "Cookie tabanlı oturum — tek klinik için yeterli", "JWT, oturum süresi, şifre sıfırlama: güvenlik protokolü; 'çalışıyor gibi görünüyor' yetmez"),
    ("Ödeme Entegrasyonu", "PayTR kodu hazır ama canlı test yok", "Para akışında BDDK uyumu ve idempotency (ödeme iki kez kesilmesin) insan gözü ister"),
    ("WhatsApp Üretimi", "Evolution API (resmi olmayan) ile pilot", "Üretim için Meta WABA API zorunlu — Evolution ile ciddi kullanımda ban riski"),
    ("Sistem Stabilitesi", "İş akışları sessizce başarısız olabiliyor", "Dead-letter queue, retry, alarm mekanizmaları yazılımcılık mimarisi ister"),
    ("Ölçeklenme", "Tek n8n instance, tek VPS", "İkinci klinik gelince webhook çakışmaları, DB sorguları elle yönetilemez"),
]
for i, (konu, durum, neden) in enumerate(limits):
    row = 24 + i
    bg = "FADBD8" if i % 2 == 0 else "FDFEFE"
    ws1.merge_cells(f"A{row}:D{row}")
    ws1[f"A{row}"] = f"  ⚠  {konu}  |  Şu an: {durum}  |  Neden şart: {neden}"
    ws1[f"A{row}"].font = normal_font(10, "5D1F1F")
    ws1[f"A{row}"].fill = fill(bg)
    ws1[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    ws1.row_dimensions[row].height = 38

ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 30
ws1.column_dimensions["D"].width = 20


# ── SAYFA 2: MVP Yol Haritası Tablosu ────────────────────────────────────────

ws2 = wb.create_sheet("MVP Yol Haritası")
ws2.sheet_view.showGridLines = False

# Başlık bandı
ws2.merge_cells("A1:E1")
ws2["A1"] = "Nefalix AI — MVP Yol Haritası Tablosu"
ws2["A1"].font = Font(bold=True, size=18, color=GOLD)
ws2["A1"].fill = fill(DARK_NAVY)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 44

# Sütun başlıkları
headers = ["Aşama / Sprint", "Yazılımcının Yapacağı İş",
           "Bu İş Ne İşe Yarıyor? (Yönetici Açıklaması)",
           "Teknik Altyapı / Araç", "Zorluk"]
for col, h in enumerate(headers, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    cell.font = Font(bold=True, size=11, color=WHITE)
    cell.fill = fill("1A3C5E")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
ws2.row_dimensions[2].height = 32

# Veri satırları
current_sprint = None
sprint_row_start = 3

for idx, (sprint, is_, aciklama, arac, zorluk) in enumerate(ROWS):
    row = idx + 3

    bg_row, font_color = SPRINT_COLORS.get(sprint, ("FFFFFF", "000000"))
    z_bg, z_font = ZORLUK_COLORS.get(zorluk, ("FFFFFF", "000000"))

    cells = [sprint, is_, aciklama, arac, zorluk]
    for col, val in enumerate(cells, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.border = thin_border()
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if col == 5:
            cell.fill = fill(z_bg)
            cell.font = Font(bold=True, size=10, color=z_font)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif col == 1:
            cell.fill = fill(bg_row)
            cell.font = Font(bold=True, size=10, color=font_color)
        else:
            cell.fill = fill(bg_row)
            cell.font = normal_font(10)
    ws2.row_dimensions[row].height = 55

# Sütun genişlikleri
col_widths = [30, 42, 48, 38, 10]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Freeze header
ws2.freeze_panes = "A3"


# ── SAYFA 3: Öncelik Özeti ────────────────────────────────────────────────────

ws3 = wb.create_sheet("Öncelik Özeti")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:C1")
ws3["A1"] = "Nefalix — Sprint Öncelik Sırası"
ws3["A1"].font = Font(bold=True, size=16, color=GOLD)
ws3["A1"].fill = fill(DARK_NAVY)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

priority_blocks = [
    ("Sprint 1-2", "Güvenlik + Auth", "1. Hafta — Bunlar olmadan ticari müşteriye açılamazsın.",
     "C8F7C5", "145A32"),
    ("Sprint 3-4", "Ödeme + WhatsApp", "İlk gelir için zorunlu iki kilit adım.",
     "FEF9E7", "7D6608"),
    ("Sprint 5-6", "Dashboard + Stabilite", "Büyüme öncesi teknik borcu kapatır.",
     "EAF6FF", "1A5276"),
    ("Sprint 7-8", "Entegrasyon + Raporlama", "Müşteri tutma ve yeni sektör açılımı.",
     "F4ECF7", "6C3483"),
    ("Gelecek Fazlar", "GEO/AEO + Otel + MCP", "Pazarlama vizyonu — şimdi geliştirme yok.",
     "FADBD8", "922B21"),
]

for i, (sprint, baslik, aciklama, bg, fc) in enumerate(priority_blocks):
    row = 3 + i * 3
    ws3.merge_cells(f"A{row}:C{row}")
    ws3[f"A{row}"] = f"  {sprint}  —  {baslik}"
    ws3[f"A{row}"].font = Font(bold=True, size=13, color=fc)
    ws3[f"A{row}"].fill = fill(bg)
    ws3[f"A{row}"].alignment = Alignment(vertical="center", indent=1)
    ws3.row_dimensions[row].height = 30

    ws3.merge_cells(f"A{row+1}:C{row+1}")
    ws3[f"A{row+1}"] = "  " + aciklama
    ws3[f"A{row+1}"].font = normal_font(11)
    ws3[f"A{row+1}"].fill = fill("FAFAFA")
    ws3[f"A{row+1}"].alignment = Alignment(vertical="center", indent=2)
    ws3.row_dimensions[row+1].height = 24

ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 50

# ── Kaydet ────────────────────────────────────────────────────────────────────
path = os.path.abspath(OUTPUT)
wb.save(path)
print(f"✅  Excel oluşturuldu: {path}")
