#!/usr/bin/env python3
"""Tek dosya — yönetici sunumu (sade dil). Google Drive'a .xlsx yükle → Sheets açılır."""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "yonetici-sunum" / "Nefalix_Yonetici_Sunum.xlsx"
TODAY = date(2026, 6, 18)

NAVY, TEAL, TEAL_L, WHITE, SLATE = "0F172A", "0D9488", "CCFBF1", "FFFFFF", "64748B"
GRAY, GREEN_L, AMBER_L, RED_L, GOLD = "F8FAFC", "DCFCE7", "FEF3C7", "FEE2E2", "CA8A04"
BORDER_C = "E2E8F0"


def border():
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def fnt(bold=False, size=11, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)


def widths(ws, w):
    for c, v in w.items():
        ws.column_dimensions[get_column_letter(c)].width = v


def title(ws, text, sub=""):
    ws.merge_cells("A1:E1")
    ws["A1"] = text
    ws["A1"].font = fnt(bold=True, size=20, color=NAVY)
    if sub:
        ws.merge_cells("A2:E2")
        ws["A2"] = sub
        ws["A2"].font = fnt(size=10, color=SLATE)


def write_sheet(ws, headers, rows, col_widths, title_text, subtitle="", header_bg=TEAL):
    ws.sheet_view.showGridLines = False
    title(ws, title_text, subtitle)
    start = 4
    for i, h in enumerate(headers, 1):
        c = ws.cell(start, i, h)
        c.font = fnt(bold=True, size=10, color=WHITE)
        c.fill = fill(header_bg)
        c.border = border()
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    for ri, row in enumerate(rows, start + 1):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.border = border()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if ri % 2 == 0:
                c.fill = fill(GRAY)
    widths(ws, col_widths)
    ws.row_dimensions[4].height = 28


def sheet_ozet(wb):
    ws = wb.active
    ws.title = "01 Özet"
    write_sheet(
        ws,
        ["Konu", "Açıklama"],
        [
            ["Bu proje ne?", "Diş kliniklerinin hasta mesajlarını, memnuniyetini ve internet yorumlarını tek yerden yönetmesini sağlayan bir program."],
            ["Kime satıyoruz?", "Diş klinikleri. İlk müşteri: Medident İstanbul."],
            ["Klinik ne kazanır?", "WhatsApp kaçmaz. Memnun hasta Google'a gider. Kötü yorum erken görülür. Sekreter işi azalır."],
            ["Biz ne satıyoruz?", "Aylık abonelik + kurulum ücreti. Programı biz işletiyoruz."],
            ["Şu an durum", "Kurulum bitti. Sunucu kimlik doğrulaması yüzünden durdu — destek açıldı."],
            ["12 ay hedef", "Medident günlük kullansın → 5–8 klinik ödesin → ayda ~60–80 bin ₺ gelir."],
            ["Yıl 1", "Yatırım yılı — kâr beklenmez. Yıl 2'de kâr hedefi."],
            ["Önerilen fiyat", "Orta paket: 7.990 ₺/ay + kurulum 14.900 ₺"],
            ["En büyük risk", "WhatsApp ve hasta verisi yasal kurallara uygun olmalı (İYS, KVKK)."],
        ],
        {1: 22, 2: 70},
        "NEFALIX — YÖNETİCİ ÖZETİ",
        f"{TODAY.strftime('%d.%m.%Y')} | Toplantıda ilk bu sayfayı açın",
    )
    ws.merge_cells("A6:B6")
    ws["A6"].fill = fill(TEAL_L)


def sheet_program(wb):
    ws = wb.create_sheet("02 Program")
    write_sheet(
        ws,
        ["Bölüm", "Klinikte ne işe yarar?", "Örnek", "Durum"],
        [
            ["Gelen kutusu", "WhatsApp mesajları tek listede. Yapay zeka taslak yazar; personel onaylar.", "İmplant fiyatı sorusu → taslak hazır", "Kurulu"],
            ["Google yorumları", "Yeni yorum → analiz + cevap taslağı. Yönetici onaylar.", "4 yıldız → nazik cevap önerisi", "Ekran gelişecek"],
            ["Hasta memnuniyeti", "Randevu sonrası puan. Memnun → Google; değilse şikâyet formu.", "9/10 → Google linki", "Kurulu"],
            ["İtibar takibi", "Şikâyetvar vb. tek yerde; aciller işaretli.", "Olumsuz haber → alarm", "Kurulu"],
            ["Geri çağırma", "Uzun süredir gelmeyene hatırlatma.", "6 ay gelmeyen hasta", "Kurulu"],
            ["Site sohbeti", "Web sitesinden yazana anında cevap.", "Gece gelen soru", "Güncellenecek"],
            ["Yönetim paneli", "Her şey tek ekran: nefalixai.com/dashboard", "Sabah 10 dk kontrol", "Hazır"],
        ],
        {1: 16, 2: 38, 3: 28, 4: 14},
        "PROGRAM NE YAPIYOR?",
        "Teknik kelime yok — klinik sahibine böyle anlatın",
    )


def sheet_yol(wb):
    ws = wb.create_sheet("03 Yol Haritası")
    write_sheet(
        ws,
        ["Dönem", "Klinik ne görür?", "Biz ne yaparız?", "Harcama", "Bitti sayılır eğer…"],
        [
            ["Haziran 2026", "Panel tekrar açılır", "Sunucuyu çalıştırırız", "Düşük", "Medident veriyi görür"],
            ["Temmuz", "WhatsApp yanıtı panelden", "Gelen kutusu iyileştirme", "Orta", "Haftada 5 gün kullanım"],
            ["Ağu–Eyl", "Google yorum onayı", "Onay ekranı", "Orta", "Yorumların yarısı panelden"],
            ["Eki–Ara", "İlk ücretli müşteriler", "Satış, sözleşme", "Yüksek", "5 klinik ödüyor"],
            ["2027 ilk yarı", "Randevu programı bağlantısı", "Klinik yazılımları entegrasyonu", "Yüksek", "10+ klinik"],
        ],
        {1: 14, 2: 28, 3: 28, 4: 10, 5: 28},
        "YOL HARİTASI",
        "Ay ay — faz kodu yok",
    )


def sheet_maliyet(wb):
    ws = wb.create_sheet("04 Maliyetler")
    write_sheet(
        ws,
        ["Gider", "Ne için?", "Ayda (₺)", "Zorunlu?"],
        [
            ["Türkiye sunucu", "Program 7/24, veri Türkiye'de", 170, "Evet"],
            ["Site barındırma", "nefalixai.com + panel", 760, "Evet"],
            ["Yapay zeka", "Taslak mesaj ve yorum analizi", 1500, "Evet"],
            ["WhatsApp resmi hat", "Yasal ticari mesaj (ileride)", 2000, "Satışta"],
            ["Hukuk / KVKK", "Sözleşme, aydınlatma", 2500, "Evet"],
            ["İYS", "Ticari mesaj izin sistemi", 500, "Satış öncesi"],
            ["Yazılımcı (yarı zaman)", "Geliştirme", 45000, "Önerilen"],
            ["Satışçı (yarı zaman)", "Demo, ziyaret", 30000, "5. müşteriden"],
            ["", "PİLOT (yazılımcı hariç)", "~6.000", ""],
            ["", "SATIŞA HAZIR (yazılımcı dahil)", "~82.000", ""],
        ],
        {1: 22, 2: 32, 3: 12, 4: 12},
        "AYLIK GİDERLER",
        "Tahmini — güncel teklif alın",
    )
    for r in (13, 14):
        ws.cell(r, 1).font = fnt(bold=True)
        ws.cell(r, 3).font = fnt(bold=True, color=TEAL)


def sheet_fiyat(wb):
    ws = wb.create_sheet("05 Fiyatlar")
    write_sheet(
        ws,
        ["Paket", "Kim alır?", "Ayda (₺)", "Kurulum (₺)", "Dahil olanlar"],
        [
            ["Başlangıç", "Tek hekim", "4.990", "9.900", "WhatsApp kutusu + site sohbeti"],
            ["Profesyonel ★", "Orta klinik", "7.990", "14.900", "Her şey: yorum, memnuniyet, itibar"],
            ["Kurumsal", "Çok şube", "14.990", "29.900", "Profesyonel + öncelikli destek"],
            ["Özel", "Zincir / hastane", "Teklif", "Teklif", "İhtiyaca göre"],
        ],
        {1: 14, 2: 16, 3: 10, 4: 12, 5: 36},
        "PAKET FİYATLARI",
        "★ = çoğu klinik için önerilen",
    )
    ws.cell(6, 1).fill = fill(GREEN_L)
    r = 10
    write_sheet_block(ws, r, ["Soru", "Cevap"], [
        ["Neden ödesin?", "1 implant randevusu 15–30 bin ₺. Program ayda 8 bin ₺."],
        ["Rakipler?", "Basit WhatsApp araçları 1.500–4.000 ₺. Biz daha kapsamlı."],
        ["Kurulum ücreti?", "İlk hafta eğitim ve bağlantı — 8–12 saat iş."],
    ], {1: 20, 2: 55})


def write_sheet_block(ws, start_row, headers, rows, col_widths):
    for i, h in enumerate(headers, 1):
        c = ws.cell(start_row, i, h)
        c.font = fnt(bold=True, size=10, color=WHITE)
        c.fill = fill(NAVY)
        c.border = border()
    for ri, row in enumerate(rows, start_row + 1):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val).border = border()
            ws.cell(ri, ci).alignment = Alignment(wrap_text=True)


def sheet_gelir(wb):
    ws = wb.create_sheet("06 Gelir")
    write_sheet(
        ws,
        ["Senaryo", "1 yıl sonu klinik", "Ort. aylık ücret", "Yıllık gelir (₺)", "Yorum"],
        [
            ["Kötü", 3, "5.990", "216.000", "Satış yavaş"],
            ["Gerçekçi ★", 8, "7.490", "718.000", "Medident referansı"],
            ["İyi", 15, "7.990", "1.440.000", "İyi satış"],
            ["Çok iyi", 25, "8.490", "2.550.000", "Hızlı büyüme"],
        ],
        {1: 12, 2: 16, 3: 14, 4: 14, 5: 22},
        "GELİR SENARYOLARI — YIL 1",
        "Yıl 1 genelde zarar — bu normal",
    )
    ws.cell(6, 1).fill = fill(GREEN_L)


def sheet_risk(wb):
    ws = wb.create_sheet("07 Riskler")
    write_sheet(
        ws,
        ["Risk", "Ne olabilir?", "Ne yapıyoruz?", "Önem"],
        [
            ["Sunucu kapalı", "Panel çalışmaz", "TR sunucu + yedek", "Yüksek"],
            ["WhatsApp kapanır", "Mesaj gitmez", "Resmi hat planı", "Yüksek"],
            ["Yasa ihlali", "Ceza", "Avukat + İYS", "Çok yüksek"],
            ["Satış olmaz", "Gelir yok", "Medident hikâyesi", "Orta"],
            ["Tek kişi", "İş durur", "Dokümantasyon", "Yüksek"],
        ],
        {1: 18, 2: 28, 3: 28, 4: 12},
        "RİSKLER",
    )
    for r in range(5, 10):
        if ws.cell(r, 4).value == "Çok yüksek":
            ws.cell(r, 4).fill = fill(RED_L)


def sheet_onay(wb):
    ws = wb.create_sheet("08 Onay Listesi")
    write_sheet(
        ws,
        ["#", "Karar", "Tutar", "ONAY\n(EVET/HAYIR)", "Not"],
        [
            [1, "Yıllık bütçe (yatırım yılı)", "~2.100.000 ₺", "", "Gelir yıl 2"],
            [2, "Hukuk / KVKK", "25.000 ₺", "", "Satış öncesi"],
            [3, "Yarı zamanlı yazılımcı", "45.000 ₺/ay", "", "Temmuz"],
            [4, "Fiyat: 7.990 ₺/ay orta paket", "—", "", ""],
            [5, "Medident referans izni", "—", "", "Case study"],
            [6, "Hedef: 5 ücretli klinik (Aralık)", "—", "", ""],
        ],
        {1: 4, 2: 36, 3: 16, 4: 14, 5: 20},
        "TOPLANTIDA ONAY",
        "Son slayt — kutuları doldurun",
    )
    for r in range(5, 11):
        ws.cell(r, 4).fill = fill(AMBER_L)


def sheet_soru(wb):
    ws = wb.create_sheet("09 Soru Cevap")
    write_sheet(
        ws,
        ["Soru", "Cevap (sunumda)"],
        [
            ["Teknik proje mi?", "Hayır — klinik işi. Randevu ve memnuniyet ön planda."],
            ["Veriler nerede?", "Türkiye sunucusunda."],
            ["Yapay zeka teşhis koyar mı?", "Hayır. Sadece randevu ve iletişim."],
            ["Ne zaman para kazanırız?", "İlk müşteri Ekim–Kasım. Anlamlı gelir 8–10 klinikte."],
            ["Rakip kim?", "Basit WhatsApp araçları ve defter. Biz hepsini tek panelde topluyoruz."],
            ["Medident durumu?", "Kurulum bitti, sunucu askıda. Açılınca eğitim."],
            ["Kaç kişi lazım?", "1 yazılımcı yarı zaman. Sonra satış + destek."],
            ["Yatırımcı gerekir mi?", "500K nakit 12–18 ay idare eder. Hız için +1–1,5M."],
        ],
        {1: 32, 2: 55},
        "SORU & CEVAP",
    )


def sheet_nasil(wb):
    ws = wb.create_sheet("10 Nasıl Yapacağız")
    write_sheet(
        ws,
        ["Adım", "Basit anlatım", "Benzetme"],
        [
            [1, "Tüm bilgi tek kasada (mesaj, yorum, puan)", "Kliniğin dijital hafızası"],
            [2, "Önce hazır araçlarla çalıştırdık", "Deneme modeli"],
            [3, "Parça parça kendi programımıza taşıyoruz", "Önce kira, sonra mülk"],
            [4, "Önce gelen kutusu ve Google yorumu", "Önce vitrin"],
            [5, "Her parça bitince klinikte test", "Pilot → ürün"],
            [6, "Eski araçlar 2027'de kapanır", "Artık kendi evimiz"],
        ],
        {1: 8, 2: 42, 3: 28},
        "NASIL GELİŞTİRECEĞİZ?",
        "Yazılım ekibine değil — yönetime anlatım",
    )


def sheet_sunum(wb):
    ws = wb.create_sheet("11 Sunum 20dk")
    write_sheet(
        ws,
        ["Dakika", "Ne söylenir?", "Sayfa"],
        [
            ["0–2", "Problem: mesaj kaçırma, geç yorum", "01 Özet"],
            ["2–5", "Çözüm: tek panel, yapay zeka, TR sunucu", "02 Program"],
            ["5–8", "Demo: nefalixai.com/dashboard", "Canlı"],
            ["8–11", "Fiyat: 7.990 ₺/ay", "05 Fiyatlar"],
            ["11–14", "Maliyet ve yıl 1 beklentisi", "04 + 06"],
            ["14–17", "Yol haritası", "03"],
            ["17–19", "Riskler ve hukuk", "07"],
            ["19–20", "Onay iste", "08 Onay"],
        ],
        {1: 10, 2: 50, 3: 14},
        "SUNUM AKIŞI — 20 DAKİKA",
        "Sunucu için rehber",
    )


def sheet_kapak(wb):
    """İlk sekme olarak taşınacak — görsel kapak."""
    ws = wb.create_sheet("SUNUM")
    ws.sheet_view.showGridLines = False
    widths(ws, {1: 4, 2: 20, 3: 20, 4: 20, 5: 20, 6: 4})
    for row in range(1, 40):
        for col in range(1, 7):
            ws.cell(row, col).fill = fill(NAVY)
    ws.merge_cells("B5:E7")
    ws["B5"] = "NEFALIX AI"
    ws["B5"].font = fnt(bold=True, size=32, color=WHITE)
    ws.merge_cells("B8:E9")
    ws["B8"] = "Diş Klinikleri İçin\nHasta İletişim & İtibar Programı"
    ws["B8"].font = fnt(size=16, color=TEAL_L)
    ws["B8"].alignment = Alignment(wrap_text=True, horizontal="center")
    ws.merge_cells("B12:E12")
    ws["B12"] = f"Yönetim Sunumu | {TODAY.strftime('%d.%m.%Y')}"
    ws["B12"].font = fnt(size=12, color=SLATE)
    boxes = [
        ("B15", "HEDEF", "8 klinik · 64.000 ₺/ay"),
        ("C15", "FİYAT", "7.990 ₺/ay"),
        ("D15", "PİLOT", "Medident İstanbul"),
        ("E15", "DURUM", "Sunucu askı — destek açık"),
    ]
    for cell, t, v in boxes:
        r = int(cell[1:])
        c0 = cell[0]
        ws.merge_cells(f"{c0}{r}:{c0}{r+2}")
        ws[cell] = f"{t}\n\n{v}"
        ws[cell].font = fnt(size=10, color=WHITE)
        ws[cell].fill = fill(TEAL)
        ws[cell].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.merge_cells("B20:E22")
    ws["B20"] = (
        "Bu dosyadaki sekmeler sırayla sunulur.\n"
        "Teknik terim yok — klinik sahibi dilinde hazırlandı.\n"
        "Google Drive: bu dosyayı yükle → Google E-Tablolar ile aç."
    )
    ws["B20"].font = fnt(size=11, color=TEAL_L)
    ws["B20"].alignment = Alignment(wrap_text=True, horizontal="center")


def main():
    wb = Workbook()
    sheet_ozet(wb)
    sheet_program(wb)
    sheet_yol(wb)
    sheet_maliyet(wb)
    sheet_fiyat(wb)
    sheet_gelir(wb)
    sheet_risk(wb)
    sheet_onay(wb)
    sheet_soru(wb)
    sheet_nasil(wb)
    sheet_sunum(wb)
    sheet_kapak(wb)

    order = [
        "SUNUM", "01 Özet", "02 Program", "03 Yol Haritası", "04 Maliyetler",
        "05 Fiyatlar", "06 Gelir", "07 Riskler", "08 Onay Listesi",
        "09 Soru Cevap", "10 Nasıl Yapacağız", "11 Sunum 20dk",
    ]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"✓ Tek dosya ({len(wb.sheetnames)} sekme): {OUT}")
    print("  Google Drive → Yükle → Sağ tık → Google E-Tablolar ile aç")


if __name__ == "__main__":
    main()
